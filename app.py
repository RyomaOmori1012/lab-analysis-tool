import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import japanize_matplotlib
from scipy import stats
from itertools import combinations
import matplotlib.transforms as transforms
import matplotlib.ticker as ticker
import itertools
from statsmodels.stats.multitest import multipletests
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import openpyxl
import re
import warnings
import traceback
import io

warnings.filterwarnings('ignore')

# ==========================================
# グローバル設定
# ==========================================
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Arial', 'MS PGothic', 'IPAexGothic']
plt.rcParams['mathtext.fontset'] = 'custom'
plt.rcParams['mathtext.rm'] = 'Arial'
plt.rcParams['mathtext.it'] = 'Arial:italic'
plt.rcParams['mathtext.bf'] = 'Arial:bold'
plt.rcParams['svg.fonttype'] = 'none'

st.set_page_config(page_title="実験データ自動解析ツール", layout="wide")
st.title("🧪 実験データ自動解析ツール")

st.markdown("""
    <style>
    textarea {
        white-space: pre !important;
        overflow-wrap: normal !important;
        overflow-x: scroll !important;
    }
    </style>
""", unsafe_allow_html=True)


# ==========================================
# 統計・計算用ヘルパー関数
# ==========================================
def calc_error(data, err_type):
    arr = np.array(data)
    arr = arr[~np.isnan(arr)]
    if len(arr) < 2: return 0.0
    sd = np.std(arr, ddof=1)
    return sd / np.sqrt(len(arr)) if "SEM" in err_type else sd

def welch_anova_games_howell(data_list):
    k = len(data_list)
    ns = np.array([len(d) for d in data_list])
    means = np.array([np.nanmean(d) for d in data_list])
    vars = np.array([np.nanvar(d, ddof=1) if len(d) > 1 else 1e-10 for d in data_list])
    vars = np.where(np.isnan(vars), 1e-10, vars)
    vars = np.where(vars <= 0, 1e-10, vars)
    
    w = ns / vars
    sum_w = np.sum(w)
    grand_mean = np.sum(w * means) / sum_w
    num = np.sum(w * (means - grand_mean)**2) / (k - 1)
    den_part = np.sum((1 - w / sum_w)**2 / (ns - 1))
    den = 1 + (2 * (k - 2) / (k**2 - 1)) * den_part
    f_val = num / den
    df1 = k - 1
    df2 = 1 / (3 / (k**2 - 1) * den_part)
    p_anova = stats.f.sf(f_val, df1, df2)
    
    pairs = []
    for i in range(k):
        for j in range(i + 1, k):
            t_val = np.abs(means[i] - means[j]) / np.sqrt(vars[i]/ns[i] + vars[j]/ns[j])
            df_num = (vars[i]/ns[i] + vars[j]/ns[j])**2
            df_den = ((vars[i]/ns[i])**2) / (ns[i]-1) + ((vars[j]/ns[j])**2) / (ns[j]-1)
            df_gh = df_num / df_den if df_den > 0 else 1e-10
            q_val = t_val * np.sqrt(2)
            try:
                p_gh = stats.studentized_range.sf(q_val, k, df_gh)
            except AttributeError:
                p_gh = stats.t.sf(t_val, df_gh) * 2 * (k * (k - 1) / 2)
                p_gh = min(p_gh, 1.0)
            pairs.append((i, j, p_gh))
            
    return p_anova, pairs

def run_statistical_test(valid_data, var_equal, is_vs_control, is_non_param, is_paired):
    k = len(valid_data)
    pairs = []
    p_anova = np.nan
    test_name = ""
    
    if k < 2: return np.nan, [], ""
        
    if k == 2:
        d1, d2 = valid_data[0], valid_data[1]
        if is_non_param:
            if is_paired:
                if len(d1) != len(d2): return np.nan, [], "Wilcoxon failed (Size mismatch)"
                try: _, p_anova = stats.wilcoxon(d1, d2); test_name = "Wilcoxon signed-rank test"
                except: p_anova = np.nan
            else:
                try: _, p_anova = stats.mannwhitneyu(d1, d2, alternative='two-sided'); test_name = "Mann-Whitney U test"
                except: p_anova = np.nan
        elif is_paired:
            if len(d1) != len(d2): return np.nan, [], "Paired t-test failed (Size mismatch)"
            try: _, p_anova = stats.ttest_rel(d1, d2); test_name = "Paired t-test"
            except: p_anova = np.nan
        else:
            try: _, p_anova = stats.ttest_ind(d1, d2, equal_var=var_equal)
            except: p_anova = np.nan
            test_name = "Student's t-test" if var_equal else "Welch's t-test"
        if not np.isnan(p_anova):
            pairs.append((0, 1, p_anova))
            
    else: 
        if is_non_param:
            try: _, p_anova = stats.kruskal(*valid_data)
            except: p_anova = np.nan
            if not np.isnan(p_anova) and p_anova < 0.05:
                raw_p, comp_pairs = [], []
                test_name = "Kruskal-Wallis test followed by Mann-Whitney U test (Holm vs Control)" if is_vs_control else "Kruskal-Wallis test followed by Mann-Whitney U test (Holm)"
                iterator = range(1, k) if is_vs_control else combinations(range(k), 2)
                for idxs in iterator:
                    i, j = (0, idxs) if is_vs_control else idxs
                    try:
                        _, p = stats.mannwhitneyu(valid_data[i], valid_data[j], alternative='two-sided')
                        raw_p.append(p); comp_pairs.append((i, j))
                    except: pass
                if raw_p:
                    _, corrected_p, _, _ = multipletests(raw_p, method='holm')
                    pairs = [(comp_pairs[m][0], comp_pairs[m][1], corrected_p[m]) for m in range(len(raw_p))]
                    
        elif is_paired:
            lens = [len(d) for d in valid_data]
            if len(set(lens)) > 1: return np.nan, [], "Friedman test failed (Size mismatch)"
            try:
                _, p_anova = stats.friedmanchisquare(*valid_data)
                test_name = "Friedman test followed by Wilcoxon signed-rank test (Holm)" if not is_vs_control else "Friedman test followed by Wilcoxon (Holm vs Control)"
            except: return np.nan, [], "Friedman test failed"
                
            if not np.isnan(p_anova) and p_anova < 0.05:
                raw_p, comp_pairs = [], []
                iterator = range(1, k) if is_vs_control else combinations(range(k), 2)
                for idxs in iterator:
                    i, j = (0, idxs) if is_vs_control else idxs
                    try:
                        _, p = stats.wilcoxon(valid_data[i], valid_data[j])
                        raw_p.append(p); comp_pairs.append((i, j))
                    except: pass
                if raw_p:
                    _, corrected_p, _, _ = multipletests(raw_p, method='holm')
                    pairs = [(comp_pairs[m][0], comp_pairs[m][1], corrected_p[m]) for m in range(len(raw_p))]
                
        else:
            if var_equal:
                test_name = "One-way ANOVA followed by Student's t-test (Holm)" if is_vs_control else "One-way ANOVA followed by Tukey's test"
                try: _, p_anova = stats.f_oneway(*valid_data)
                except: p_anova = np.nan
                
                if not np.isnan(p_anova) and p_anova < 0.05:
                    if is_vs_control:
                        raw_p, comp_pairs = [], []
                        for j in range(1, k):
                            try:
                                _, p = stats.ttest_ind(valid_data[0], valid_data[j], equal_var=True)
                                raw_p.append(p); comp_pairs.append((0, j))
                            except: pass
                        if raw_p:
                            _, corrected_p, _, _ = multipletests(raw_p, method='holm')
                            pairs = [(comp_pairs[m][0], comp_pairs[m][1], corrected_p[m]) for m in range(len(raw_p))]
                    else:
                        all_v, all_g = [], []
                        for p_idx, d in enumerate(valid_data):
                            all_v.extend(d)
                            all_g.extend([p_idx] * len(d))
                        try:
                            tukey = pairwise_tukeyhsd(all_v, all_g, alpha=0.05)
                            df_t = pd.DataFrame(data=tukey._results_table.data[1:], columns=tukey._results_table.data[0])
                            for _, row in df_t.iterrows():
                                pairs.append((int(row['group1']), int(row['group2']), row['p-adj']))
                        except: pass
            else:
                test_name = "Welch's ANOVA followed by Welch's t-test (Holm)" if is_vs_control else "Welch's ANOVA followed by Games-Howell test"
                try: p_anova, gh_pairs = welch_anova_games_howell(valid_data)
                except: p_anova = np.nan
                
                if not np.isnan(p_anova) and p_anova < 0.05:
                    if is_vs_control:
                        raw_p, comp_pairs = [], []
                        for j in range(1, k):
                            try:
                                _, p = stats.ttest_ind(valid_data[0], valid_data[j], equal_var=False)
                                raw_p.append(p); comp_pairs.append((0, j))
                            except: pass
                        if raw_p:
                            _, corrected_p, _, _ = multipletests(raw_p, method='holm')
                            pairs = [(comp_pairs[m][0], comp_pairs[m][1], corrected_p[m]) for m in range(len(raw_p))]
                    else:
                        pairs = gh_pairs
                        
    return p_anova, pairs, test_name

# ==========================================
# パーサー関数
# ==========================================
def parse_text(text):
    if not text.strip(): return [np.nan]
    res = []
    for line in text.replace(',', '\n').split('\n'):
        if line.strip():
            try: res.append(float(line.strip()))
            except ValueError: res.append(np.nan)
    return res if res else [np.nan]

def parse_plate(text):
    if not text.strip(): return np.full((8, 12), np.nan)
    lines = [line for line in text.replace('\r', '').split('\n')]
    data = []
    for line in lines:
        if not line.strip(): continue
        row = []
        parts = line.split('\t') if '\t' in line else re.sub(r'[\s,]+', ',', line.strip()).split(',')
        for x in parts:
            x = x.strip()
            if not x: row.append(np.nan)
            else:
                try: row.append(float(x))
                except ValueError: row.append(np.nan)
        while len(row) < 12: row.append(np.nan)
        data.append(row[:12])
    while len(data) < 8: data.append([np.nan] * 12)
    return np.array(data[:8])

def parse_idx(text, is_alpha=False):
    res = []
    try:
        for p in text.replace(' ', '').split(','):
            if not p: continue
            if '-' in p:
                start, end = p.split('-')
                if start and end: 
                    res.extend(range(ord(start.upper())-65, ord(end.upper())-65+1) if is_alpha else range(int(start)-1, int(end)))
            else: 
                res.append(ord(p.upper())-65 if is_alpha else int(p)-1)
    except Exception: pass 
    return list(set(res))


# ==========================================
# UI 描画・解析用関数群
# ==========================================

def render_mtt_analysis(input_data, config):
    i_rows, i_cols = parse_idx(config['mtt_ignore_row'], True), parse_idx(config['mtt_ignore_col'], False)
    b_cols, c_cols, s_cols = parse_idx(config['mtt_blank_col'], False), parse_idx(config['mtt_control_col'], False), parse_idx(config['mtt_sample_cols'], False)
    s_cols.sort()
    valid_rows = [r for r in range(8) if r not in i_rows]
    
    safe_dilution = config['mtt_dilution'] if config['mtt_dilution'] != 0 else 1.0
    conc_vals_plot = [config['mtt_start_conc'] / (safe_dilution ** i) for i in range(len(s_cols))][::-1]
    s_cols_plot = s_cols[::-1] if "左が高濃度" in config['mtt_conc_direction'] else s_cols
    
    plates_data, plate_names, ctrl_err_pct_list = [], [], []
    for idx, item in enumerate(input_data):
        pn, pd_text = item[0], item[1]
        arr = parse_plate(pd_text); plate_names.append(pn or f"Plate {idx+1}")
        blank_vals = [arr[r, c] for r in valid_rows for c in b_cols if c not in i_cols and not np.isnan(arr[r, c])]
        blank_mean = np.nanmean(blank_vals) if blank_vals else 0.0
        
        ctrl_vals = [arr[r, c] - blank_mean for r in valid_rows for c in c_cols if c not in i_cols and not np.isnan(arr[r, c])]
        ctrl_mean = np.nanmean(ctrl_vals) if ctrl_vals else np.nan
        c_err = calc_error(ctrl_vals, config['error_bar_type'])
        ctrl_err_pct_list.append((c_err / ctrl_mean) * 100 if not np.isnan(ctrl_mean) and ctrl_mean != 0 else 0)
        
        if np.isnan(ctrl_mean) or ctrl_mean == 0: plates_data.append(np.full((8, 12), np.nan))
        else: plates_data.append((arr - blank_mean) / ctrl_mean * 100)
    
    num_p = len(plates_data)
    indiv_figs = []
    
    for i in range(num_p):
        fig_i, ax_i = plt.subplots(figsize=(6, 4))
        fig_i.patch.set_facecolor('white')
        ax_i.set_facecolor('white')
        
        means_i = [np.nanmean(plates_data[i][valid_rows, c]) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        errs_i = [calc_error(plates_data[i][valid_rows, c], config['error_bar_type']) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        
        ax_i.errorbar(conc_vals_plot, means_i, yerr=errs_i, fmt='-o', color='black', capsize=4, mfc='black', mec='black', lw=1.5)
        ax_i.set_xscale('log')
        
        mtt_max_y_i = 125.0
        for m, e in zip(means_i, errs_i):
            if not np.isnan(m) and not np.isnan(e):
                mtt_max_y_i = max(mtt_max_y_i, (m + e) * 1.15)
        ax_i.set_ylim(bottom=0, top=mtt_max_y_i)
        ax_i.yaxis.set_major_locator(ticker.MultipleLocator(20))
        
        for spine in ax_i.spines.values(): spine.set_color('black'); spine.set_linewidth(1.2)
        ax_i.minorticks_off()
        
        if config['mtt_custom_xticks'].strip() and len(conc_vals_plot) > 0:
            try:
                c_ticks = [float(x.strip()) for x in config['mtt_custom_xticks'].split(',') if x.strip()]
                all_x = conc_vals_plot + c_ticks
                min_x, max_x = min(all_x), max(all_x)
                low_exp, high_exp = int(np.floor(np.log10(min_x))), int(np.ceil(np.log10(max_x)))
                default_ticks = [10**e for e in range(low_exp, high_exp + 1)]
                combined_ticks = sorted(list(set(default_ticks + c_ticks)))
                ax_i.set_xticks(combined_ticks)
                ax_i.get_xaxis().set_major_formatter(ticker.ScalarFormatter())
                ax_i.set_xlim(min_x * 0.8, max_x * 1.2)
            except Exception: pass
        else:
            ax_i.xaxis.set_major_formatter(ticker.FuncFormatter(lambda y, _: '{:g}'.format(y)))
            
        ax_i.tick_params(direction='in', length=5, width=1.2, labelsize=12, colors='black', which='major')
        ax_i.set_ylabel(config['ylabel_input'], fontsize=14, fontweight='bold', fontname='Arial', labelpad=8)
        ax_i.set_xlabel(f"{config['l_name']} [{config['mtt_unit']}]", fontsize=14, fontweight='bold', fontname='Arial', labelpad=8)
        n_indiv = max([np.count_nonzero(~np.isnan(plates_data[i][valid_rows, c])) for c in s_cols_plot]) if s_cols_plot else len(valid_rows)
        ax_i.set_title(f"n={n_indiv}", fontsize=14, pad=15, loc='right')
        indiv_figs.append((plate_names[i], fig_i))

    fig_comb, ax = plt.subplots(figsize=(7, 5))
    fig_comb.patch.set_facecolor('white')
    ax.set_facecolor('white')
    
    mtt_max_y_comb = 125.0
    colors = sns.color_palette("Set1", max(num_p, 2)) if num_p > 1 else ['black']
    for i in range(num_p):
        means = [np.nanmean(plates_data[i][valid_rows, c]) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        errs = [calc_error(plates_data[i][valid_rows, c], config['error_bar_type']) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        ax.plot(conc_vals_plot, means, '-o', color=colors[i], mfc=colors[i], mec=colors[i], lw=1.8, label=plate_names[i])
        ax.errorbar(conc_vals_plot, means, yerr=errs, fmt='none', color=colors[i], capsize=4, lw=1.8)
        
        for m, e in zip(means, errs):
            if not np.isnan(m) and not np.isnan(e): mtt_max_y_comb = max(mtt_max_y_comb, (m + e) * 1.15)

    plotted_stars, mtt_test_name = set(), ""
    dropped_warnings, non_param_warnings = set(), set()
    
    for idx_c, c in enumerate(s_cols_plot):
        col_data = [d[~np.isnan(d)] for d in [plates_data[p][valid_rows, c] for p in range(num_p)]]
        col_data_valid = []
        for p_idx, d in enumerate(col_data):
            if len(d) >= 2: col_data_valid.append(d)
            else: dropped_warnings.add(f"{plate_names[p_idx]} ({conc_vals_plot[idx_c]} {config['mtt_unit']})")
        
        if config['is_non_param'] and any(len(d) <= 3 for d in col_data_valid):
            non_param_warnings.add("MTTデータ")
        
        p_anova, pairs, t_name = run_statistical_test(col_data_valid, config['var_equal'], config['is_vs_control'], config['is_non_param'], config['is_paired'])
        if t_name: mtt_test_name = t_name
        
        min_p = min([p_val for _, _, p_val in pairs]) if pairs else np.nan

        if not np.isnan(min_p) and min_p < 0.05:
            stars = "***" if min_p < 0.001 else "**" if min_p < 0.01 else "*"
            plotted_stars.add(stars)
            max_mean_err_c = max([np.nanmean(d) + calc_error(d, config['error_bar_type']) for d in col_data_valid if not np.isnan(np.nanmean(d)) and not np.isnan(calc_error(d, config['error_bar_type']))] + [0])
            text_y = max_mean_err_c + (mtt_max_y_comb * 0.05)
            mtt_max_y_comb = max(mtt_max_y_comb, text_y * 1.15)
            ax.text(conc_vals_plot[idx_c], text_y, stars, ha='center', va='bottom', fontsize=14, fontweight='bold', color='black')

    if dropped_warnings: st.warning(f"⚠️ データ不足により除外: {', '.join(dropped_warnings)}")
    if non_param_warnings: st.info("💡 n≤3の場合、ノンパラメトリック検定で有意差が出ない可能性があります。")

    ax.set_xscale('log')
    ax.set_ylim(bottom=0, top=mtt_max_y_comb)
    ax.yaxis.set_major_locator(ticker.MultipleLocator(20))
    for spine in ax.spines.values(): spine.set_color('black'); spine.set_linewidth(1.2)
    ax.minorticks_off()
    
    if config['mtt_custom_xticks'].strip() and len(conc_vals_plot) > 0:
        try:
            c_ticks = [float(x.strip()) for x in config['mtt_custom_xticks'].split(',') if x.strip()]
            all_x = conc_vals_plot + c_ticks
            min_x, max_x = min(all_x), max(all_x)
            low_exp, high_exp = int(np.floor(np.log10(min_x))), int(np.ceil(np.log10(max_x)))
            combined_ticks = sorted(list(set([10**e for e in range(low_exp, high_exp + 1)] + c_ticks)))
            ax.set_xticks(combined_ticks)
            ax.get_xaxis().set_major_formatter(ticker.ScalarFormatter())
            ax.set_xlim(min_x * 0.8, max_x * 1.2)
        except Exception: pass
    else: ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda y, _: '{:g}'.format(y)))
        
    ax.tick_params(direction='in', length=5, width=1.2, labelsize=12, colors='black', which='major')
    ax.set_ylabel(config['ylabel_input'], fontsize=14, fontweight='bold', labelpad=8)
    ax.set_xlabel(f"{config['l_name']} [{config['mtt_unit']}]", fontsize=14, fontweight='bold', labelpad=8)
    if num_p > 1: ax.legend(loc='lower left', frameon=False, prop={'size': 13})
    
    max_n = max([np.count_nonzero(~np.isnan(plates_data[i][valid_rows, c])) for i in range(num_p) for c in s_cols_plot]) if num_p > 0 else 0
    star_str = ", " + ", ".join([f"{s} p < {0.05 if s=='*' else 0.01 if s=='**' else 0.001}" for s in ["*", "**", "***"] if s in plotted_stars]) if plotted_stars else ""
    ax.set_title(f"{mtt_test_name}{star_str}, n={max_n}" if mtt_test_name and num_p > 1 else f"n={max_n}", fontsize=14, pad=15, loc='right')

    st.pyplot(fig_comb)
    
    # --- Excel 書き出し ---
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        mtt_summary_dict = {"濃度 (Concentration)": [0.0] + [float(x) for x in conc_vals_plot]}
        err_label = "SEM(%)" if "SEM" in config['error_bar_type'] else "SD(%)"
        for i, p_name in enumerate(plate_names):
            mtt_summary_dict[f"{p_name}_Mean(%)"] = [100.0] + [float(np.nanmean(plates_data[i][valid_rows, c])) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
            mtt_summary_dict[f"{p_name}_{err_label}"] = [float(ctrl_err_pct_list[i])] + [float(calc_error(plates_data[i][valid_rows, c], config['error_bar_type'])) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        pd.DataFrame(mtt_summary_dict).to_excel(writer, sheet_name='Summary', index=False)
        
        ws = writer.book['Summary']
        ws.cell(row=2, column=len(mtt_summary_dict) + 2, value="💡 【エラーバー付き折れ線グラフの最短作成手順】")
        ws.cell(row=3, column=len(mtt_summary_dict) + 2, value="1. 『濃度』の列と、グラフにしたい『〇〇_Mean(%)』の列を同時選択し、[挿入] ＞ [散布図 (直線とマーカー)] を作成。")
        ws.cell(row=4, column=len(mtt_summary_dict) + 2, value="2. 作成されたグラフの横軸をクリックして[軸の書式設定]を開き、『対数目盛を表示する』にチェック。")
        ws.cell(row=5, column=len(mtt_summary_dict) + 2, value="3. グラフのプロット線をクリックし、[＋] ＞ [誤差範囲] ＞ [その他の誤差範囲オプション]。")
        ws.cell(row=6, column=len(mtt_summary_dict) + 2, value="4. 『カスタム』にチェックを入れ、『値の指定』をクリック。")
        ws.cell(row=7, column=len(mtt_summary_dict) + 2, value=f"5. 正負両方の選択ボックスに、対応する『〇〇_{err_label}』の数値をドラッグして指定すれば完成！")

        long_mtt_list = []
        for i, p_name in enumerate(plate_names):
            ctrl_vals = [plates_data[i][r, c] for r in valid_rows for c in c_cols if c not in i_cols]
            for val in ctrl_vals:
                if not np.isnan(val): long_mtt_list.append({"条件名": f"{p_name}_0_{config['mtt_unit']}", "正規化生存率 (%)": float(val)})
            for idx_c, c in enumerate(s_cols_plot):
                for val in plates_data[i][valid_rows, c]:
                    if not np.isnan(val): long_mtt_list.append({"条件名": f"{p_name}_{conc_vals_plot[idx_c]}_{config['mtt_unit']}", "正規化生存率 (%)": float(val)})
        pd.DataFrame(long_mtt_list).to_excel(writer, sheet_name='Normalized_Data', index=False)
        
        for i in range(num_p):
            df_norm = pd.DataFrame(plates_data[i])
            df_norm.index = ['A','B','C','D','E','F','G','H']
            df_norm.columns = [str(x+1) for x in range(12)]
            df_norm.to_excel(writer, sheet_name=re.sub(r'[\\/*?:\[\]]', '', f"Plate_{i+1}_{plate_names[i]}")[:31])

    st.download_button("📥 Excelデータをダウンロード (全データ・統計詳細シート同梱)", excel_buffer.getvalue(), "Analysis_Data.xlsx", type="primary", use_container_width=True)
    dl_col1, dl_col2 = st.columns(2)
    buf_c = io.BytesIO()
    fig_comb.savefig(buf_c, format='svg', bbox_inches='tight')
    with dl_col1: st.download_button("📥 統合グラフ(SVG)を保存", buf_c.getvalue(), "Combined_Graph.svg", "image/svg+xml", use_container_width=True)
    with st.expander("個別プレートのグラフ(SVG)をダウンロード"):
        for p_name, f in indiv_figs:
            buf_i = io.BytesIO()
            f.savefig(buf_i, format='svg', bbox_inches='tight')
            st.download_button(f"📥 {p_name} のグラフ", buf_i.getvalue(), f"{p_name}_Graph.svg", "image/svg+xml")

def render_single_target(input_data, config):
    upper_labels, lower_labels, internal_ids, raw_processed = [], [], [], {}
    dropped_warnings, non_param_warnings = set(), set()
    
    for idx, item in enumerate(input_data):
        u, d, val_t_list, val_l_list = item[0], item[1], item[2], item[3] if len(item) > 3 else []
        if config['is_microscope']: raw_processed[f"C_{idx}"] = parse_text(val_t_list[0])
        else:
            t_nums, l_nums = parse_text(val_t_list[0]), parse_text(val_l_list[0])
            length = max(len(t_nums), len(l_nums))
            t_nums.extend([np.nan] * (length - len(t_nums)))
            l_nums.extend([np.nan] * (length - len(l_nums)))
            
            processed = []
            for t, l in zip(t_nums, l_nums):
                if np.isnan(t) or np.isnan(l): processed.append(np.nan)
                elif config['is_qpcr']: processed.append(t - l)
                else: processed.append(np.nan if l == 0 else t / l)
            raw_processed[f"C_{idx}"] = processed
        
        upper_labels.append(u or f"U_{idx+1}")
        lower_labels.append(d or "")
        internal_ids.append(f"C_{idx}")
    
    if not any(len([v for v in raw_processed[uid] if not np.isnan(v)]) > 0 for uid in internal_ids): st.stop()
        
    final_norm = {}
    ctrl_id = internal_ids[0]
    for i, uid in enumerate(internal_ids):
        c_id = internal_ids[lower_labels.index(lower_labels[i])] if 'グループ' in config['norm_mode'] else ctrl_id
        c_mean = np.nanmean(raw_processed[c_id])
        if np.isnan(c_mean) or c_mean == 0: c_mean = 1.0 
        
        if config['is_qpcr']: final_norm[uid] = [2 ** -(v - c_mean) for v in raw_processed[uid]]
        else: final_norm[uid] = [v / c_mean for v in raw_processed[uid]]
    
    p_pairs, test_desc_flat = [], ""
    groupings = [[u for u in internal_ids if lower_labels[internal_ids.index(u)] == low] for low in sorted(list(set(lower_labels)), key=lambda x: lower_labels.index(x))] if config['is_grouped_test'] else [internal_ids]

    for grp in groupings:
        valid_uids, valid_data = [], []
        for u in grp:
            non_nan_data = [v for v in raw_processed[u] if not np.isnan(v)]
            if len(non_nan_data) >= 2:
                valid_uids.append(u); valid_data.append(non_nan_data)
            else:
                uid_idx = internal_ids.index(u)
                dropped_warnings.add(f"{upper_labels[uid_idx]} ({lower_labels[uid_idx]})" if lower_labels[uid_idx] else upper_labels[uid_idx])
        
        if len(valid_data) < 2: continue
        if config['is_non_param'] and any(len(d) <= 3 for d in valid_data): non_param_warnings.add("解析データ")
        
        _, pairs, t_name = run_statistical_test(valid_data, config['var_equal'], config['is_vs_control'], config['is_non_param'], config['is_paired'])
        if t_name: test_desc_flat = t_name
        for i_idx, j_idx, p_val in pairs: p_pairs.append((valid_uids[i_idx], valid_uids[j_idx], p_val))

    if dropped_warnings: st.warning(f"⚠️ データ不足により除外: {', '.join(dropped_warnings)}")
    if non_param_warnings: st.info("💡 n≤3の場合、ノンパラメトリック検定で有意差が出ない可能性があります。")

    unique_low = sorted(list(set(lower_labels)), key=lambda x: lower_labels.index(x))
    unique_up = sorted(list(set(upper_labels)), key=lambda x: upper_labels.index(x))
    gray_palette = ['black', 'darkgray', 'lightgray', 'dimgray', 'whitesmoke', '#E0E0E0']
    palette = {u: gray_palette[i % len(gray_palette)] for i, u in enumerate(unique_up)} if "色分け" in config['color_mode'] else {u: "black" for u in unique_up}
    
    fig, ax = plt.subplots(figsize=(max(4.0, len(internal_ids)*1.5+1.5), 5.5))
    fig.patch.set_facecolor('white'); ax.set_facecolor('white')
    
    x_coords, bar_width = {}, config['bar_width']
    if config['layout_mode'] == "条件ごとにグループ化":
        current_x = 0
        for low in unique_low:
            members = [i for i, l in enumerate(lower_labels) if l == low]
            for i in members:
                x_coords[internal_ids[i]] = current_x
                current_x += bar_width + 0.02
            current_x += 0.5
    else:
        for i, uid in enumerate(internal_ids): x_coords[uid] = float(i)

    if config['is_microscope']:
        box_data_safe = [[v for v in final_norm[uid] if not np.isnan(v)] or [np.nan] for uid in internal_ids]
        ax.boxplot(box_data_safe, positions=[x_coords[uid] for uid in internal_ids], widths=bar_width*1.5, patch_artist=True, 
                   boxprops=dict(facecolor='white', color='black', linewidth=1.2), 
                   capprops=dict(color='black', linewidth=1.2), whiskerprops=dict(color='black', linewidth=1.2),
                   medianprops=dict(color='black', linewidth=1.5), flierprops=dict(marker='o', markerfacecolor='black', markeredgecolor='black', alpha=0.8, markersize=4))
    else:
        for i, uid in enumerate(internal_ids):
            mean_val, err_val = np.nanmean(final_norm[uid]), calc_error(final_norm[uid], config['error_bar_type'])
            ax.bar(x_coords[uid], mean_val if not np.isnan(mean_val) else 0, yerr=err_val if not np.isnan(err_val) else 0, 
                   width=bar_width, color=palette[upper_labels[i]], edgecolor='black', capsize=3, error_kw=dict(ecolor='black', lw=1.2), 
                   label=upper_labels[i] if i == upper_labels.index(upper_labels[i]) else "")

    for spine in ax.spines.values(): spine.set_visible(True); spine.set_color('black'); spine.set_linewidth(1.5)
    ax.tick_params(axis='y', colors='black', direction='in', left=True, right=False, length=5, width=1.5, labelsize=14)
    ax.tick_params(axis='x', bottom=False, top=False); ax.set_xticklabels([]) 
    trans = transforms.blended_transform_factory(ax.transData, ax.transAxes)
    
    if config['layout_mode'] == "条件ごとにグループ化" and "色分け" in config['color_mode']:
        for low in unique_low:
            xs = [x_coords[internal_ids[i]] for i, l in enumerate(lower_labels) if l == low]
            if xs: ax.text(sum(xs) / len(xs), -0.05, low, ha='center', va='top', transform=trans, fontsize=16, fontweight='bold', color='black')
    else:
        for i, uid in enumerate(internal_ids):
            ax.text(x_coords[uid], -0.05, upper_labels[i], ha='center', va='top', transform=trans, fontsize=16, color='black', fontweight='bold')
        for label, elements in [(k, list(g)) for k, g in itertools.groupby(enumerate(lower_labels), key=lambda x: x[1])]:
            if not label: continue
            xs = [x_coords[internal_ids[x[0]]] for x in elements]
            x_start, x_end = min(xs), max(xs)
            if x_start != x_end: ax.plot([x_start - bar_width/2, x_end + bar_width/2], [-0.16, -0.16], color='black', lw=1.5, transform=trans, clip_on=False)
            ax.text((x_start + x_end) / 2, -0.21, label, ha='center', va='top', transform=trans, fontsize=16, fontweight='bold', color='black')

    all_vals = [v for vals in final_norm.values() for v in vals if not np.isnan(v)]
    current_max_y = max(all_vals + [0]) if all_vals else 1.0
    for uid in internal_ids:
        m, e = np.nanmean(final_norm[uid]), calc_error(final_norm[uid], config['error_bar_type'])
        if not np.isnan(m) and not np.isnan(e): current_max_y = max(current_max_y, m + e)
    if current_max_y == 0: current_max_y = 1.0
    
    y_shift, h, base_bracket_y, max_element_y = current_max_y * 0.15, current_max_y * 0.025, current_max_y * 1.10, current_max_y
    levels, max_level, sig_pairs, plotted_stars = [], 0, [], set()
    
    for u1, u2, p in p_pairs:
        if p >= 0.05 or np.isnan(p) or u1 not in x_coords or u2 not in x_coords: continue
        sig_pairs.append((min(x_coords[u1], x_coords[u2]), max(x_coords[u1], x_coords[u2]), "***" if p < 0.001 else "**" if p < 0.01 else "*"))
    
    for x_start, x_end, stars in sorted(sig_pairs, key=lambda x: x[1] - x[0]):
        plotted_stars.add(stars)
        placed_level = next((l_idx for l_idx, intervals in enumerate(levels) if not any(not (x_end < s or x_start > e) for s, e in intervals)), -1)
        if placed_level == -1: placed_level = len(levels); levels.append([])
        levels[placed_level].append((x_start, x_end)); max_level = max(max_level, placed_level)
        by = base_bracket_y + placed_level * y_shift; max_element_y = max(max_element_y, by + h)
        ax.plot([x_start, x_start, x_end, x_end], [by - h, by, by, by - h], color='black', lw=1.2)
        ax.text((x_start + x_end) / 2, by + h*0.2, stars, ha='center', va='bottom', color='black', fontsize=14, fontweight='bold')

    ax.set_ylim(0, max(current_max_y * 1.2, max_element_y * 1.15))
    x_vals = list(x_coords.values())
    if x_vals: ax.set_xlim(min(x_vals) - 0.6, max(x_vals) + 0.6)
    ax.set_ylabel(config['ylabel_input'], fontsize=16, fontweight="bold", color='black', labelpad=10)
    
    if "色分け" in config['color_mode']:
        # ★ バグ修正箇所: 凡例の重複排除を安全な方法に変更
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        if by_label: ax.legend(by_label.values(), by_label.keys(), loc='upper right', frameon=False, prop={'size': 12, 'weight': 'bold'})

    n_list = [len([v for v in raw_processed[u] if not np.isnan(v)]) for u in internal_ids]
    expected_n = n_list[0] if n_list and len(set(n_list)) == 1 else "varies"
    star_str = ", " + ", ".join([f"{s} p < {0.05 if s=='*' else 0.01 if s=='**' else 0.001}" for s in ["*", "**", "***"] if s in plotted_stars]) if plotted_stars else ""
    title_str = f"{test_desc_flat}{star_str}" if config['is_microscope'] else f"{test_desc_flat}{star_str}, n={expected_n}" if test_desc_flat else f"n={expected_n}"
    if title_str: ax.set_title(title_str, fontsize=14, pad=15, loc='right')

    st.pyplot(fig)
    
    # --- Excel 書き出し ---
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        err_label = "SEM" if "SEM" in config['error_bar_type'] else "SD"
        
        summary_df = pd.DataFrame({'上段ラベル': upper_labels, '下段ラベル': lower_labels, '平均': [np.nanmean(final_norm[u]) for u in internal_ids], err_label: [calc_error(final_norm[u], config['error_bar_type']) for u in internal_ids]})
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        pd.DataFrame([{"条件名": f"{upper_labels[i]} ({lower_labels[i]})" if lower_labels[i] else upper_labels[i], "正規化データ": float(val)} for i, u in enumerate(internal_ids) for val in final_norm[u] if not np.isnan(val)]).to_excel(writer, sheet_name='Normalized_Data', index=False)
        
        stat_data = []
        for u1, u2, p in p_pairs:
            idx1, idx2 = internal_ids.index(u1), internal_ids.index(u2)
            c1 = f"{upper_labels[idx1]} ({lower_labels[idx1]})" if lower_labels[idx1] else upper_labels[idx1]
            c2 = f"{upper_labels[idx2]} ({lower_labels[idx2]})" if lower_labels[idx2] else upper_labels[idx2]
            stat_data.append({"比較": f"{c1} vs {c2}", "p値": p if not np.isnan(p) else "N/A", "判定": "***" if p<0.001 else "**" if p<0.01 else "*" if p<0.05 else "ns" if not np.isnan(p) else "N/A"})
        if stat_data: pd.DataFrame(stat_data).to_excel(writer, sheet_name='Statistical_Details', index=False)

        try:
            if config['is_microscope']:
                ws = writer.book['Normalized_Data']
                ws.cell(row=2, column=4, value="💡 【箱ひげ図の最短作成手順】")
                ws.cell(row=3, column=4, value="1. 左のA列とB列をすべて全選択します。")
                ws.cell(row=4, column=4, value="2. [挿入]タブ ＞ [統計グラフ] ＞ [箱ひげ図] をクリックします。")
            elif config['layout_mode'] == "条件ごとにグループ化":
                matrix_mean = pd.DataFrame(index=unique_up, columns=unique_low)
                matrix_sd = pd.DataFrame(index=unique_up, columns=unique_low)
                for i, uid in enumerate(internal_ids):
                    matrix_mean.at[upper_labels[i], lower_labels[i]] = np.nanmean(final_norm[uid])
                    matrix_sd.at[upper_labels[i], lower_labels[i]] = calc_error(final_norm[uid], config['error_bar_type'])
                
                matrix_mean.to_excel(writer, sheet_name='Summary_Matrix', startrow=1, startcol=0)
                matrix_sd.to_excel(writer, sheet_name='Summary_Matrix', startrow=len(unique_up)+4, startcol=0)
                
                ws = writer.book['Summary_Matrix']
                ws.cell(row=1, column=1, value="【平均値 (Mean)】")
                ws.cell(row=len(unique_up)+4, column=1, value=f"【{err_label}】")
                
                sc = len(unique_low) + 3
                ws.cell(row=2, column=sc, value="💡 【グループ化棒グラフの最短作成手順】")
                ws.cell(row=3, column=sc, value="1. 左上の【平均値】の表(A2から)を丸ごと選択し、[挿入] ＞ [2D 縦棒 (集合縦棒)] をクリック。")
                ws.cell(row=4, column=sc, value="2. 追加された棒をクリックし、[誤差範囲] ＞ [その他の誤差範囲オプション] ＞ [カスタム]")
                ws.cell(row=5, column=sc, value=f"3. 値の指定で、下の【{err_label}】の表の該当する行をドラッグして指定すれば完成です！")
            else:
                ws = writer.book['Summary']
                sc = len(summary_df.columns) + 2
                ws.cell(row=2, column=sc, value="💡 【エラーバー付き棒グラフの最短作成手順】")
                ws.cell(row=3, column=sc, value="1. 左の『上段ラベル』と『平均』の列を選択し、[挿入] ＞ [縦棒グラフ] を作成。")
                ws.cell(row=4, column=sc, value="2. グラフの棒をクリックし、[＋] ＞ [誤差範囲] ＞ [その他の誤差範囲オプション]。")
                ws.cell(row=5, column=sc, value="3. 『カスタム』にチェックを入れ、『値の指定』。")
                ws.cell(row=6, column=sc, value=f"4. 正負両方に、左の『{err_label}』の数値をドラッグして指定すれば完成！")
        except Exception:
            pass

    col_dl1, col_dl2 = st.columns(2)
    buf_svg = io.BytesIO()
    fig.savefig(buf_svg, format='svg', bbox_inches='tight')
    with col_dl1: st.download_button("📥 Excelデータをダウンロード", excel_buffer.getvalue(), "Analysis_Data.xlsx", type="primary", use_container_width=True)
    with col_dl2: st.download_button("📥 完成グラフ(SVG)を保存", buf_svg.getvalue(), "Graph.svg", "image/svg+xml", use_container_width=True)

def render_multi_target(input_data, config):
    upper_labels, lower_labels, internal_ids = [], [], []
    raw_processed_multi = {j: {} for j in range(config['num_targets'])}
    dropped_warnings, non_param_warnings = set(), set()
    
    for idx, item in enumerate(input_data):
        u, d, val_t_list, val_l_list = item[0], item[1], item[2], item[3] if len(item) > 3 else []
        for j in range(config['num_targets']):
            if config['is_microscope']: raw_processed_multi[j][f"C_{idx}"] = parse_text(val_t_list[j])
            else:
                t_nums, l_nums = parse_text(val_t_list[j]), parse_text(val_l_list[j])
                length = max(len(t_nums), len(l_nums))
                t_nums_ext = t_nums + [np.nan] * (length - len(t_nums))
                l_nums_ext = l_nums + [np.nan] * (length - len(l_nums))
                
                processed = []
                for t, l in zip(t_nums_ext, l_nums_ext):
                    if np.isnan(t) or np.isnan(l): processed.append(np.nan)
                    elif config['is_qpcr']: processed.append(t - l)
                    else: processed.append(np.nan if l == 0 else t / l)
                raw_processed_multi[j][f"C_{idx}"] = processed
                
        upper_labels.append(u or f"U_{idx+1}")
        lower_labels.append(d or "")
        internal_ids.append(f"C_{idx}")
        
    if not any(len([v for v in raw_processed_multi[0][uid] if not np.isnan(v)]) > 0 for uid in internal_ids): st.stop()
    
    final_norm_multi = {j: {} for j in range(config['num_targets'])}
    ctrl_id = internal_ids[0]
    
    for j in range(config['num_targets']):
        for i, uid in enumerate(internal_ids):
            c_id = internal_ids[lower_labels.index(lower_labels[i])] if 'グループ' in config['norm_mode'] else ctrl_id
            c_mean = np.nanmean(raw_processed_multi[j][c_id])
            if np.isnan(c_mean) or c_mean == 0: c_mean = 1.0 
            if config['is_qpcr']: final_norm_multi[j][uid] = [2 ** -(v - c_mean) for v in raw_processed_multi[j][uid]]
            else: final_norm_multi[j][uid] = [v / c_mean for v in raw_processed_multi[j][uid]]
    
    p_pairs_multi = {j: [] for j in range(config['num_targets'])}
    test_desc_flat = ""
    groupings = [[u for u in internal_ids if lower_labels[internal_ids.index(u)] == low] for low in sorted(list(set(lower_labels)), key=lambda x: lower_labels.index(x))] if config['is_grouped_test'] else [internal_ids]

    for j in range(config['num_targets']):
        for grp in groupings:
            valid_uids, valid_data = [], []
            for u in grp:
                non_nan_data = [v for v in raw_processed_multi[j][u] if not np.isnan(v)]
                if len(non_nan_data) >= 2:
                    valid_uids.append(u); valid_data.append(non_nan_data)
                else:
                    uid_idx = internal_ids.index(u)
                    dropped_warnings.add(f"{config['target_names'][j]}の{upper_labels[uid_idx]}")
            
            if len(valid_data) < 2: continue
            if config['is_non_param'] and any(len(d) <= 3 for d in valid_data): non_param_warnings.add("解析データ")
                
            _, pairs, t_name = run_statistical_test(valid_data, config['var_equal'], config['is_vs_control'], config['is_non_param'], config['is_paired'])
            if t_name: test_desc_flat = t_name
            for i_idx, j_idx, p_val in pairs: p_pairs_multi[j].append((valid_uids[i_idx], valid_uids[j_idx], p_val))

    if dropped_warnings: st.warning(f"⚠️ データ不足により除外: {', '.join(dropped_warnings)}")
    if non_param_warnings: st.info("💡 n≤3の場合、ノンパラメトリック検定で有意差が出ない可能性があります。")

    unique_up = sorted(list(set(upper_labels)), key=lambda x: upper_labels.index(x))
    gray_palette = ['black', 'darkgray', 'lightgray', 'dimgray', 'whitesmoke', '#E0E0E0']
    if "色分け" in config['color_mode']:
        colors = sns.color_palette("Set1", max(len(unique_up), 2))
        palette = {u: colors[i % len(colors)] for i, u in enumerate(unique_up)}
    else:
        palette = {u: gray_palette[i % len(gray_palette)] for i, u in enumerate(unique_up)}
        
    fig, ax = plt.subplots(figsize=(max(6.0, config['num_targets'] * len(unique_up) * 1.0), 5.5))
    fig.patch.set_facecolor('white'); ax.set_facecolor('white')
    
    bar_width, x_coords_multi, target_centers, current_x = config['bar_width'], {j: {} for j in range(config['num_targets'])}, [], 0
    
    for j in range(config['num_targets']):
        g_start = current_x
        for i, uid in enumerate(internal_ids):
            x_coords_multi[j][uid] = current_x
            if config['is_microscope']:
                d_list_safe = [v for v in final_norm_multi[j][uid] if not np.isnan(v)] or [np.nan]
                ax.boxplot([d_list_safe], positions=[current_x], widths=bar_width*1.5, patch_artist=True, 
                           boxprops=dict(facecolor='white', color='black', linewidth=1.2), capprops=dict(color='black', linewidth=1.2),
                           whiskerprops=dict(color='black', linewidth=1.2), medianprops=dict(color='black', linewidth=1.5), 
                           flierprops=dict(marker='o', markerfacecolor='black', markeredgecolor='black', alpha=0.8, markersize=4))
            else:
                mean_val, err_val = np.nanmean(final_norm_multi[j][uid]), calc_error(final_norm_multi[j][uid], config['error_bar_type'])
                ax.bar(current_x, mean_val if not np.isnan(mean_val) else 0, yerr=err_val if not np.isnan(err_val) else 0, 
                       width=bar_width, color=palette[upper_labels[i]], edgecolor='black', capsize=3, error_kw=dict(ecolor='black', lw=1.2), 
                       label=upper_labels[i] if (j == 0 and i == upper_labels.index(upper_labels[i])) else "")
            current_x += bar_width + 0.02
        target_centers.append((g_start + current_x - bar_width - 0.02) / 2)
        current_x += 0.8

    for spine in ax.spines.values(): spine.set_visible(True); spine.set_color('black'); spine.set_linewidth(1.5)
    ax.tick_params(axis='y', colors='black', direction='in', left=True, right=False, length=5, width=1.5, labelsize=14)
    ax.tick_params(axis='x', bottom=False, top=False)
    ax.set_xticks(target_centers); ax.set_xticklabels(config['target_names'], fontsize=16, fontweight='bold', color='black')
    
    all_vals = [v for j in range(config['num_targets']) for vals in final_norm_multi[j].values() for v in vals if not np.isnan(v)]
    current_max_y = max(all_vals + [0]) if all_vals else 1.0
    for j in range(config['num_targets']):
        for uid in internal_ids:
            m, e = np.nanmean(final_norm_multi[j][uid]), calc_error(final_norm_multi[j][uid], config['error_bar_type'])
            if not np.isnan(m) and not np.isnan(e): current_max_y = max(current_max_y, m + e)
    if current_max_y == 0: current_max_y = 1.0
    
    y_shift, h, base_bracket_y, max_element_y = current_max_y * 0.15, current_max_y * 0.025, current_max_y * 1.10, current_max_y
    plotted_stars = set()
    
    for j in range(config['num_targets']):
        levels, max_level, sig_pairs = [], 0, []
        for u1, u2, p in p_pairs_multi[j]:
            if p >= 0.05 or np.isnan(p) or u1 not in x_coords_multi[j] or u2 not in x_coords_multi[j]: continue
            sig_pairs.append((min(x_coords_multi[j][u1], x_coords_multi[j][u2]), max(x_coords_multi[j][u1], x_coords_multi[j][u2]), "***" if p < 0.001 else "**" if p < 0.01 else "*"))
        
        for x_start, x_end, stars in sorted(sig_pairs, key=lambda x: x[1] - x[0]):
            plotted_stars.add(stars)
            placed_level = next((l_idx for l_idx, intervals in enumerate(levels) if not any(not (x_end < s or x_start > e) for s, e in intervals)), -1)
            if placed_level == -1: placed_level = len(levels); levels.append([])
            levels[placed_level].append((x_start, x_end)); max_level = max(max_level, placed_level)
            by = base_bracket_y + placed_level * y_shift; max_element_y = max(max_element_y, by + h)
            ax.plot([x_start, x_start, x_end, x_end], [by - h, by, by, by - h], color='black', lw=1.2)
            ax.text((x_start + x_end) / 2, by + h*0.2, stars, ha='center', va='bottom', color='black', fontsize=14, fontweight='bold')
        base_bracket_y += (max_level + 1) * y_shift if sig_pairs else 0
        max_element_y = max(max_element_y, base_bracket_y)

    ax.set_ylim(0, max(current_max_y * 1.2, max_element_y * 1.15))
    ax.set_xlim(-0.6, current_x - 0.8 + 0.6)
    ax.set_ylabel(config['ylabel_input'], fontsize=16, fontweight="bold", color='black', labelpad=10)
    
    if "色分け" in config['color_mode']:
        # ★ バグ修正箇所: 凡例の重複排除を安全な方法に変更
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        if by_label: ax.legend(by_label.values(), by_label.keys(), loc='upper right', frameon=False, prop={'size': 12, 'weight': 'bold'})

    expected_n = n_list[0] if (n_list := [len([v for v in raw_processed_multi[0][u] if not np.isnan(v)]) for u in internal_ids]) and len(set(n_list)) == 1 else "varies"
    star_str = ", " + ", ".join([f"{s} p < {0.05 if s=='*' else 0.01 if s=='**' else 0.001}" for s in ["*", "**", "***"] if s in plotted_stars]) if plotted_stars else ""
    title_str = f"{test_desc_flat}{star_str}, n={expected_n}" if test_desc_flat else f"n={expected_n}"
    ax.set_title(title_str, fontsize=14, pad=15, loc='right')

    st.pyplot(fig)
    
    # --- Excel 書き出し ---
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        err_label = "SEM" if "SEM" in config['error_bar_type'] else "SD"
        pd.DataFrame([{'ターゲット名': config['target_names'][j], '上段ラベル': upper_labels[i], '下段ラベル': lower_labels[i], '平均': np.nanmean(final_norm_multi[j][u]), err_label: calc_error(final_norm_multi[j][u], config['error_bar_type'])} for j in range(config['num_targets']) for i, u in enumerate(internal_ids)]).to_excel(writer, sheet_name='Summary', index=False)
        
        ws = writer.book['Summary']
        ws.cell(row=2, column=7, value="💡 【複数ターゲットの棒グラフ最短作成手順】")
        ws.cell(row=3, column=7, value="1. 1行目（見出し）を選択し、[データ]タブ ＞ [フィルター] をクリック。")
        ws.cell(row=4, column=7, value="2. 『ターゲット名』の▼をクリックし、グラフにしたいターゲットを1つだけ選んで絞り込む。")
        ws.cell(row=5, column=7, value="3. 表示された『上段ラベル』と『平均』の列を同時選択し、[挿入] ＞ [縦棒グラフ] を作成。")
        ws.cell(row=6, column=7, value="4. グラフの棒をクリックし、[＋] ＞ [誤差範囲] ＞ [その他の誤差範囲オプション]。")
        ws.cell(row=7, column=7, value=f"5. 『カスタム』を選び、『値の指定』で正負両方に、フィルター後の表示されている『{err_label}』の列をドラッグして指定すれば完成！")

        pd.DataFrame([{"ターゲット名": config['target_names'][j], "条件名": f"{upper_labels[i]} ({lower_labels[i]})" if lower_labels[i] else upper_labels[i], "正規化データ": float(val)} for j in range(config['num_targets']) for i, u in enumerate(internal_ids) for val in final_norm_multi[j][u] if not np.isnan(val)]).to_excel(writer, sheet_name='Normalized_Data', index=False)
        
        stat_data = []
        for j in range(config['num_targets']):
            for u1, u2, p in p_pairs_multi[j]:
                idx1, idx2 = internal_ids.index(u1), internal_ids.index(u2)
                c1 = f"{upper_labels[idx1]} ({lower_labels[idx1]})" if lower_labels[idx1] else upper_labels[idx1]
                c2 = f"{upper_labels[idx2]} ({lower_labels[idx2]})" if lower_labels[idx2] else upper_labels[idx2]
                stat_data.append({"ターゲット名": config['target_names'][j], "比較": f"{c1} vs {c2}", "p値": p if not np.isnan(p) else "N/A", "判定": "***" if p<0.001 else "**" if p<0.01 else "*" if p<0.05 else "ns" if not np.isnan(p) else "N/A"})
        if stat_data: pd.DataFrame(stat_data).to_excel(writer, sheet_name='Statistical_Details', index=False)

    col_dl1, col_dl2 = st.columns(2)
    buf_svg = io.BytesIO()
    fig.savefig(buf_svg, format='svg', bbox_inches='tight')
    with col_dl1: st.download_button("📥 Excelデータをダウンロード", excel_buffer.getvalue(), "Analysis_Data.xlsx", type="primary", use_container_width=True)
    with col_dl2: st.download_button("📥 完成グラフ(SVG)を保存", buf_svg.getvalue(), "Graph.svg", "image/svg+xml", use_container_width=True)


# ==========================================
# メイン実行関数
# ==========================================
def main():
    # --- サイドバー設定 ---
    st.sidebar.header("⚙️ 全体設定")
    selected_exp = st.sidebar.selectbox('実験手法:', ['Western Blotting (WB)', 'HPLC', 'qPCR', 'MTT Assay (細胞生存率)', '蛍光顕微鏡 (Box Plot)'])
    num_cond = st.sidebar.number_input('手動モード時の条件数:', min_value=1, max_value=20, value=2, step=1)

    is_mtt = 'MTT' in selected_exp
    is_microscope = '顕微鏡' in selected_exp
    is_qpcr = 'qPCR' in selected_exp
    is_hplc = 'HPLC' in selected_exp
    is_multi_capable = 'WB' in selected_exp or is_qpcr or is_hplc

    num_targets = st.sidebar.number_input('ターゲットの数 (1つのグラフにまとめる数):', min_value=1, max_value=10, value=1, step=1) if is_multi_capable else 1

    if 'WB' in selected_exp:
        t_label, t_ph, l_label, l_ph, y_label_def = 'Target:', '例: HO-1', 'Loading Control:', '例: HSP90', 'Relative Band Intensity'
    elif 'HPLC' in selected_exp:
        t_label, t_ph, l_label, l_ph, y_label_def = '物質名:', '例: PpIX', 'タンパク質濃度:', '例: protein', 'Intracellular Concentration\n[nmol / mg ・ protein]'
    elif 'qPCR' in selected_exp:
        t_label, t_ph, l_label, l_ph, y_label_def = 'Target:', '例: PDK1', 'Loading Control:', '例: β-ACTIN', 'Relative mRNA level'
    elif is_mtt:
        t_label, t_ph, l_label, l_ph, y_label_def = '細胞株:', '例: PC3', '薬剤名:', '例: ALA', 'Cell Viability [%]'
    elif is_microscope:
        t_label, t_ph, l_label, l_ph, y_label_def = '観察対象:', '例: ROS / GFP', '', '', 'Relative Fluorescence Intensity'

    is_common_loading = True
    target_names, loading_names = [], []

    if num_targets == 1:
        c_side1, c_side2 = st.sidebar.columns(2)
        with c_side1: t_name_raw = st.text_input(t_label, placeholder=t_ph).strip()
        with c_side2: l_name_raw = st.text_input(l_label, placeholder=l_ph).strip() if not is_microscope else ""
        target_names.append(t_name_raw or ("Cell Line" if is_mtt else "Target"))
        loading_names.append(l_name_raw or ("Drug" if is_mtt else ("" if is_microscope else "Loading Control")))
    else:
        if not is_mtt and not is_microscope:
            is_common_loading = "共通" in st.sidebar.radio("Loading Controlの扱い:", ["共通 (全てのターゲットで同じデータを使用)", "ターゲットごとに個別"])
            if is_common_loading:
                l_name_raw = st.sidebar.text_input(f'共通の {l_label}', placeholder=l_ph).strip()
                loading_names = [l_name_raw or "Loading Control"] * num_targets
            
        st.sidebar.markdown("**ターゲット設定**")
        for i in range(num_targets):
            target_names.append(st.sidebar.text_input(f'{t_label} {i+1}:', placeholder=f'Target {i+1}').strip() or f"Target {i+1}")
            if not is_common_loading and not is_mtt and not is_microscope:
                loading_names.append(st.sidebar.text_input(f'{l_label} {i+1}:', placeholder=f'Loading {i+1}').strip() or f"Loading {i+1}")
                st.sidebar.markdown("---")

    t_name = target_names[0]
    l_name = loading_names[0] if loading_names else ("Drug" if is_mtt else "" if is_microscope else "Loading Control")
    y_label_full = y_label_def if (is_mtt or is_microscope or is_hplc or num_targets > 1) else f"{y_label_def}\n[{t_name} / {l_name}]"
    ylabel_input = st.sidebar.text_area('Y軸ラベル:', value=y_label_full, height=68)

    st.sidebar.markdown("---")
    st.sidebar.header("🖌️ レイアウト・統計設定")

    error_bar_type = st.sidebar.radio("エラーバーの種類:", ["SD (標準偏差)", "SEM (標準誤差)"])

    if not is_mtt:
        layout_mode = "条件ごとにグループ化" if num_targets > 1 else st.sidebar.radio("棒の配置:", ["均等に並べる", "条件ごとにグループ化"])
        if num_targets > 1: st.sidebar.info("💡 複数ターゲットモードでは、棒の配置は自動的に「ターゲット毎のグループ化」になります。")
            
        color_mode = st.sidebar.radio("配色:", ["すべて黒", "上段ラベルで色分け（黒/グレー）"])
        bar_width_input = st.sidebar.slider("棒の太さ調整:", min_value=0.05, max_value=0.80, value=(0.25 if layout_mode == "条件ごとにグループ化" else 0.17), step=0.01)
        
        pairing_options = ['独立 (パラメトリック)', '独立 (ノンパラメトリック)'] if is_microscope else ['独立 (パラメトリック)', '独立 (ノンパラメトリック)', '対応あり (パラメトリック)', '対応あり (ノンパラメトリック)']
        pairing_mode = st.sidebar.radio('統計検定の前提:', pairing_options)
        
        var_equal = '等しい' in st.sidebar.radio('ばらつき(分散)の仮定:', ['分散が等しいと仮定する (古典的)', '分散が異なると仮定する (Welch等)']) if ('パラメトリック' in pairing_mode and '独立' in pairing_mode) else False
            
        is_vs_control = 'Control' in st.sidebar.radio('比較方式 (3条件以上の場合):', ['すべての組み合わせを総当たりで比較', '一番左の群(Control)とだけ比較'])
        is_non_param = 'ノンパラメトリック' in pairing_mode
        is_paired = '対応あり' in pairing_mode
        norm_mode = st.sidebar.radio('規格化:', ['全体基準 (一番上の条件で全て規格化)', 'グループ基準 (下段ラベル毎の先頭条件で規格化)'])
        is_grouped_test = ('グループ内' in st.sidebar.radio('検定範囲:', ['すべての条件間で検定', 'グループ内でのみ検定 (下段ラベルが同じ条件間)'])) if num_targets == 1 else True
    else:
        layout_mode, color_mode, norm_mode = "", "", ""
        bar_width_input = 0.17
        pairing_mode = st.sidebar.radio('統計検定の前提:', ['独立 (パラメトリック)', '独立 (ノンパラメトリック)', '対応あり (パラメトリック)', '対応あり (ノンパラメトリック)'])
        var_equal = '等しい' in st.sidebar.radio('ばらつき(分散)の仮定:', ['分散が等しいと仮定する (古典的)', '分散が異なると仮定する (Welch等)']) if ('パラメトリック' in pairing_mode and '独立' in pairing_mode) else False
        is_vs_control = 'Control' in st.sidebar.radio('比較方式 (3条件以上の場合):', ['すべての組み合わせを総当たりで比較', '一番左の群(Control)とだけ比較'])
        is_non_param = 'ノンパラメトリック' in pairing_mode
        is_paired = '対応あり' in pairing_mode
        is_grouped_test = False

    u_label_name = "系列名" if (not is_mtt and layout_mode == "条件ごとにグループ化" and "色分け" in color_mode) else ("横ラベル上段" if not is_mtt else "")
    d_label_name = "横ラベル" if (not is_mtt and layout_mode == "条件ごとにグループ化" and "色分け" in color_mode) else ("横ラベル下段" if not is_mtt else "")
    paste_t_label = 'Target' if 'WB' in selected_exp or 'qPCR' in selected_exp else ('物質名' if 'HPLC' in selected_exp else t_name)
    paste_l_label = 'Loading Control' if 'WB' in selected_exp or 'qPCR' in selected_exp else ('タンパク質濃度' if 'HPLC' in selected_exp else l_name)

    st.markdown("---")
    
    # configディクショナリの構築（関数に渡すため）
    config = {
        'is_mtt': is_mtt, 'is_microscope': is_microscope, 'is_qpcr': is_qpcr, 'is_hplc': is_hplc,
        'num_targets': num_targets, 'target_names': target_names, 'loading_names': loading_names,
        't_name': t_name, 'l_name': l_name, 'ylabel_input': ylabel_input,
        'error_bar_type': error_bar_type, 'layout_mode': layout_mode, 'color_mode': color_mode,
        'bar_width': bar_width_input, 'var_equal': var_equal, 'is_vs_control': is_vs_control,
        'is_non_param': is_non_param, 'is_paired': is_paired, 'norm_mode': norm_mode,
        'is_grouped_test': is_grouped_test, 'u_label_name': u_label_name, 'd_label_name': d_label_name,
        'paste_t_label': paste_t_label, 'paste_l_label': paste_l_label
    }

    # --- UI 入力画面 ---
    col_input, col_graph = st.columns([1.2, 1.0], gap="large")
    input_data = []

    with col_input:
        st.header("📝 データ入力")
        if is_mtt:
            c1, c2, c3 = st.columns(3)
            config['mtt_ignore_row'] = c1.text_input('空のWell(除外行):', 'A, H')
            config['mtt_ignore_col'] = c2.text_input('空のWell(除外列):', '1')
            config['mtt_blank_col'] = c3.text_input('バックグラウンド（培地のみ）(列):', '12')
            c4, c5 = st.columns(2)
            config['mtt_control_col'] = c4.text_input('Control（細胞生存率100%の基準）(列):', '11')
            config['mtt_sample_cols'] = c5.text_input('Sample(列):', '2-10')
            c6, c7, c8 = st.columns(3)
            config['mtt_start_conc'] = c6.number_input('開始濃度:', value=4000.0)
            config['mtt_dilution'] = c7.number_input('希釈倍率(n倍):', value=2.0)
            config['mtt_unit'] = c8.text_input('単位:', 'μM')
            config['mtt_conc_direction'] = st.radio("濃度の配置方向:", ["左が高濃度 (右へ希釈)", "右が高濃度 (左へ希釈)"], horizontal=True)
            config['mtt_custom_xticks'] = st.text_input('横軸の目盛りに明示したい数値（カンマ区切りで追加指定、空欄なら自動）', value='', placeholder='例: 10, 50, 250')
            
            for i in range(num_cond):
                p_name = st.text_input(f'プレート {i+1} 条件名:', placeholder=f'例: プレート{i+1}', key=f"pname_{i}")
                p_data = st.text_area(f'プレート {i+1} データ (8行x12列):', placeholder='ここにペースト', height=220, key=f"pdata_{i}")
                input_data.append((p_name, p_data))
        else:
            input_mode = "手動で1条件ずつ入力" if is_microscope else st.radio("入力モード:", ["エクセル列ごとに一括ペースト（おすすめ✨）", "手動で1条件ずつ入力"], horizontal=True)
            
            if input_mode == "エクセル列ごとに一括ペースト（おすすめ✨）":
                st.info("💡 エクセル上で離れた列にあってもOK！必要な列だけを個別にコピーしてペーストしてください。\nペースト後に出現する表で、離れたサンプルを隣同士に整理できます。")
                
                bulk_n, bulk_t_list, bulk_l_list = "", [], []
                if num_targets == 1:
                    c_n, c_l, c_t = st.columns(3)
                    bulk_n = c_n.text_area("1. 【名前】の列をペースト", height=150, placeholder="例:\nsiNC\nsiHSPA9")
                    bulk_l_list = [c_l.text_area(f"2. 【{paste_l_label}】", height=150)]
                    bulk_t_list = [c_t.text_area(f"3. 【{paste_t_label}】", height=150)]
                else:
                    if is_common_loading:
                        cols_bulk = st.columns(num_targets + 2)
                        bulk_n = cols_bulk[0].text_area("1. 【名前】", height=150)
                        bulk_l_list = [cols_bulk[1].text_area(f"2. 共通【{paste_l_label}】", height=150)] * num_targets
                        for j in range(num_targets): bulk_t_list.append(cols_bulk[j+2].text_area(f"{j+3}. 【{target_names[j]}】", height=150))
                    else:
                        bulk_n = st.columns([1, 3])[0].text_area("1. 【名前】列", height=150)
                        for j in range(num_targets):
                            ct, cl = st.columns(2)
                            bulk_t_list.append(ct.text_area(f"【{target_names[j]}】", height=150, key=f"bulk_t_{j}"))
                            bulk_l_list.append(cl.text_area(f"対応する【{loading_names[j]}】", height=150, key=f"bulk_l_{j}"))
                
                if bulk_n.strip():
                    try:
                        n_lines = [line.strip() for line in bulk_n.replace('\r', '').split('\n') if line.strip()]
                        t_lines_list = [[line.strip() for line in b.replace('\r', '').split('\n') if line.strip()] if b.strip() else [] for b in bulk_t_list]
                        l_lines_list = [[line.strip() for line in b.replace('\r', '').split('\n') if line.strip()] if b.strip() else [] for b in bulk_l_list]

                        raw_dict = {}
                        for i, name in enumerate(n_lines):
                            if not name: continue
                            if name not in raw_dict: raw_dict[name] = {'t': [[] for _ in range(num_targets)], 'l': [[] for _ in range(num_targets)]}
                            for j in range(num_targets):
                                if i < len(t_lines_list[j]): raw_dict[name]['t'][j].extend([float(x) for x in re.sub(r'[\s,]+', ',', t_lines_list[j][i]).split(',') if x.strip()])
                                if i < len(l_lines_list[j]): raw_dict[name]['l'][j].extend([float(x) for x in re.sub(r'[\s,]+', ',', l_lines_list[j][i]).split(',') if x.strip()])

                        mapping_df = pd.DataFrame({"表示順 (1,2,3...)": range(1, len(raw_dict) + 1), "エクセルの名前 (読取専用)": list(raw_dict.keys()), u_label_name: list(raw_dict.keys()), f"{d_label_name} (空欄可)": [""] * len(raw_dict)})
                        edited_df = st.data_editor(mapping_df, hide_index=True, use_container_width=True, disabled=["エクセルの名前 (読取専用)"]).sort_values(by="表示順 (1,2,3...)")
                        
                        for _, row in edited_df.iterrows():
                            orig_name = row["エクセルの名前 (読取専用)"]
                            u_label = str(row[u_label_name]).strip() if pd.notna(row[u_label_name]) else ""
                            d_label = str(row[f"{d_label_name} (空欄可)"]).strip() if pd.notna(row[f"{d_label_name} (空欄可)"]) else ""
                            input_data.append((u_label, d_label, ['\n'.join(map(str, raw_dict[orig_name]['t'][j])) for j in range(num_targets)], ['\n'.join(map(str, raw_dict[orig_name]['l'][j])) for j in range(num_targets)]))
                    except Exception: st.error("データの読み取りに失敗しました。数字や文字の形式を確認してください。")
            else:
                for i in range(num_cond):
                    st.markdown(f"**条件 {i+1}**") if num_targets > 1 else None
                    if is_microscope:
                        col_up, col_dn, col_t = st.columns([1, 1, 3.0]) if num_targets == 1 else (*st.columns(2), None)
                        n_up = (col_up if num_targets == 1 else st.columns(2)[0]).text_input(f'{u_label_name}:', placeholder='Control' if i==0 else f'Cond_{i+1}', key=f"up_{i}")
                        n_down = (col_dn if num_targets == 1 else st.columns(2)[1]).text_input(f'{d_label_name}:', placeholder='(空欄可)', key=f"dn_{i}")
                        
                        if num_targets == 1:
                            input_data.append((n_up, n_down, [col_t.text_area(f'{paste_t_label}:', placeholder='縦にペースト', height=100, key=f"t_{i}")], []))
                        else:
                            cols_manual = st.columns(num_targets)
                            input_data.append((n_up, n_down, [cols_manual[j].text_area(f'{target_names[j]}:', placeholder='縦にペースト', height=100, key=f"t_{i}_{j}") for j in range(num_targets)], []))
                    elif num_targets == 1:
                        col_up, col_dn, col_l, col_t = st.columns([1, 1, 1.5, 1.5])
                        input_data.append((col_up.text_input(f'{u_label_name}:', placeholder='Control' if i==0 else f'Cond_{i+1}', key=f"up_{i}"), col_dn.text_input(f'{d_label_name}:', placeholder='(空欄可)', key=f"dn_{i}"), [col_t.text_area(f'{paste_t_label}:', placeholder='縦にペースト', height=100, key=f"t_{i}")], [col_l.text_area(f'{paste_l_label}:', placeholder='縦にペースト', height=100, key=f"l_{i}")]))
                    else:
                        col_up, col_dn = st.columns(2)
                        n_up = col_up.text_input(f'{u_label_name}:', placeholder='Control' if i==0 else f'Cond_{i+1}', key=f"up_{i}")
                        n_down = col_dn.text_input(f'{d_label_name}:', placeholder='(空欄可)', key=f"dn_{i}")
                        if is_common_loading:
                            cols_manual = st.columns([1.5] + [1.5]*num_targets)
                            n_l_list = [cols_manual[0].text_area(f'共通の {paste_l_label}:', placeholder='縦にペースト', height=100, key=f"l_{i}")] * num_targets
                            input_data.append((n_up, n_down, [cols_manual[1+j].text_area(f'{target_names[j]}:', placeholder='縦にペースト', height=100, key=f"t_{i}_{j}") for j in range(num_targets)], n_l_list))
                        else:
                            n_t_list, n_l_list = [], []
                            for j in range(num_targets):
                                ct, cl = st.columns(2)
                                n_t_list.append(ct.text_area(f'{target_names[j]}:', placeholder='縦にペースト', height=100, key=f"t_{i}_{j}"))
                                n_l_list.append(cl.text_area(f'対応する {loading_names[j]}:', placeholder='縦にペースト', height=100, key=f"l_{i}_{j}"))
                            input_data.append((n_up, n_down, n_t_list, n_l_list))

    # --- UI グラフ表示部 ---
    with col_graph:
        st.header("📊 リアルタイムプレビュー")
        st.info("💡 左の枠に文字を打つとグラフの枠が連動し、数値をペーストすると棒が出現します。")
        
        try:
            if config['is_mtt']:
                render_mtt_analysis(input_data, config)
            elif config['num_targets'] == 1:
                render_single_target(input_data, config)
            else:
                render_multi_target(input_data, config)
        except Exception as e:
            st.error(f"グラフ描画中にエラーが発生しました: {e}")
            st.code(traceback.format_exc())

if __name__ == "__main__":
    main()
