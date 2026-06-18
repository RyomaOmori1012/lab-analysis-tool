import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import japanize_matplotlib
import matplotlib.transforms as transforms
import matplotlib.ticker as ticker
import itertools
import openpyxl
import re
import io
import warnings

warnings.filterwarnings('ignore')

# 先ほど作成した utils.py から計算用関数を読み込む
from utils import calc_error, run_statistical_test, parse_text, parse_plate, parse_idx

# --- フォントのグローバル設定 ---
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Arial', 'MS PGothic', 'IPAexGothic']
plt.rcParams['mathtext.fontset'] = 'custom'
plt.rcParams['mathtext.rm'] = 'Arial'
plt.rcParams['mathtext.it'] = 'Arial:italic'
plt.rcParams['mathtext.bf'] = 'Arial:bold'
plt.rcParams['svg.fonttype'] = 'none'

def render_mtt_analysis(input_data, config):
    i_rows, i_cols = parse_idx(config['mtt_ignore_row'], True), parse_idx(config['mtt_ignore_col'], False)
    b_cols, c_cols, s_cols = parse_idx(config['mtt_blank_col'], False), parse_idx(config['mtt_control_col'], False), parse_idx(config['mtt_sample_cols'], False)
    s_cols.sort()
    valid_rows = [r for r in range(8) if r not in i_rows]
    
    safe_dilution = config['mtt_dilution'] if config['mtt_dilution'] != 0 else 1.0
    conc_vals_plot = [config['mtt_start_conc'] / (safe_dilution ** i) for i in range(len(s_cols))][::-1]
    s_cols_plot = s_cols[::-1] if "左が高濃度" in config['mtt_conc_direction'] else s_cols
    
    plates_data, plate_names, ctrl_err_pct_list = [], [], []
    exclude_flags = [] # ★追加
    
    for idx, item in enumerate(input_data):
        pn, pd_text = item[0], item[1]
        exclude_flag = item[2] if len(item) > 2 else False # ★追加
        exclude_flags.append(exclude_flag)
        
        arr = parse_plate(pd_text); plate_names.append(pn or f"Plate {idx+1}")
        blank_vals = [arr[r, c] for r in valid_rows for c in b_cols if c not in i_cols and not np.isnan(arr[r, c])]
        blank_mean = np.nanmean(blank_vals) if blank_vals else 0.0
        
        ctrl_vals = [arr[r, c] - blank_mean for r in valid_rows for c in c_cols if c not in i_cols and not np.isnan(arr[r, c])]
        ctrl_mean = np.nanmean(ctrl_vals) if ctrl_vals else np.nan
        c_err = calc_error(ctrl_vals, config['error_bar_type'])
        ctrl_err_pct_list.append((c_err / ctrl_mean) * 100 if not np.isnan(ctrl_mean) and ctrl_mean != 0 else 0)
        
        if np.isnan(ctrl_mean) or ctrl_mean == 0: plates_data.append(np.full((8, 12), np.nan))
        else: plates_data.append((arr - blank_mean) / ctrl_mean * 100)
    
    num_p = len(plates_data)
    indiv_figs = []
    
    for i in range(num_p):
        fig_i, ax_i = plt.subplots(figsize=(6, 4))
        fig_i.patch.set_facecolor('white')
        ax_i.set_facecolor('white')
        
        means_i = [np.nanmean(plates_data[i][valid_rows, c]) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        errs_i = [calc_error(plates_data[i][valid_rows, c], config['error_bar_type']) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        
        ax_i.errorbar(conc_vals_plot, means_i, yerr=errs_i, fmt='-o', color='black', capsize=4, mfc='black', mec='black', lw=1.5)
        ax_i.set_xscale('log')
        
        mtt_max_y_i = 125.0
        for m, e in zip(means_i, errs_i):
            if not np.isnan(m) and not np.isnan(e):
                mtt_max_y_i = max(mtt_max_y_i, (m + e) * 1.15)
        ax_i.set_ylim(bottom=0, top=mtt_max_y_i)
        ax_i.yaxis.set_major_locator(ticker.MultipleLocator(20))
        
        for spine in ax_i.spines.values(): spine.set_color('black'); spine.set_linewidth(1.2)
        ax_i.minorticks_off()
        
        if config['mtt_custom_xticks'].strip() and len(conc_vals_plot) > 0:
            try:
                c_ticks = [float(x.strip()) for x in config['mtt_custom_xticks'].split(',') if x.strip()]
                all_x = conc_vals_plot + c_ticks
                min_x, max_x = min(all_x), max(all_x)
                low_exp, high_exp = int(np.floor(np.log10(min_x))), int(np.ceil(np.log10(max_x)))
                default_ticks = [10**e for e in range(low_exp, high_exp + 1)]
                combined_ticks = sorted(list(set(default_ticks + c_ticks)))
                ax_i.set_xticks(combined_ticks)
                ax_i.get_xaxis().set_major_formatter(ticker.ScalarFormatter())
                ax_i.set_xlim(min_x * 0.8, max_x * 1.2)
            except Exception: pass
        else:
            ax_i.xaxis.set_major_formatter(ticker.FuncFormatter(lambda y, _: '{:g}'.format(y)))
            
        ax_i.tick_params(direction='in', length=5, width=1.2, labelsize=12, colors='black', which='major')
        ax_i.set_ylabel(config['ylabel_input'], fontsize=14, fontweight='bold', fontname='Arial', labelpad=8)
        ax_i.set_xlabel(f"{config['l_name']} [{config['mtt_unit']}]", fontsize=14, fontweight='bold', fontname='Arial', labelpad=8)
        n_indiv = max([np.count_nonzero(~np.isnan(plates_data[i][valid_rows, c])) for c in s_cols_plot]) if s_cols_plot else len(valid_rows)
        ax_i.set_title(f"n={n_indiv}", fontsize=14, pad=15, loc='right')
        indiv_figs.append((plate_names[i], fig_i))

    fig_comb, ax = plt.subplots(figsize=(7, 5))
    fig_comb.patch.set_facecolor('white')
    ax.set_facecolor('white')
    
    mtt_max_y_comb = 125.0
    colors = sns.color_palette("Set1", max(num_p, 2)) if num_p > 1 else ['black']
    for i in range(num_p):
        means = [np.nanmean(plates_data[i][valid_rows, c]) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        errs = [calc_error(plates_data[i][valid_rows, c], config['error_bar_type']) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        ax.plot(conc_vals_plot, means, '-o', color=colors[i], mfc=colors[i], mec=colors[i], lw=1.8, label=plate_names[i])
        ax.errorbar(conc_vals_plot, means, yerr=errs, fmt='none', color=colors[i], capsize=4, lw=1.8)
        
        for m, e in zip(means, errs):
            if not np.isnan(m) and not np.isnan(e): mtt_max_y_comb = max(mtt_max_y_comb, (m + e) * 1.15)

    plotted_stars, mtt_test_name = set(), ""
    dropped_warnings, non_param_warnings = set(), set()
    
    for idx_c, c in enumerate(s_cols_plot):
        col_data_valid = []
        for p_idx in range(num_p):
            if exclude_flags[p_idx]: continue # ★除外処理
            d = plates_data[p_idx][valid_rows, c]
            d_clean = d[~np.isnan(d)]
            if len(d_clean) >= 2: col_data_valid.append(d_clean)
            else: dropped_warnings.add(f"{plate_names[p_idx]} ({conc_vals_plot[idx_c]} {config['mtt_unit']})")
        
        if config['is_non_param'] and any(len(d) <= 3 for d in col_data_valid):
            non_param_warnings.add("MTTデータ")
        
        min_p = np.nan
        # ★トグルONの時だけ検定
        if config.get('show_stats', True) and len(col_data_valid) >= 2:
            p_anova, pairs, t_name = run_statistical_test(col_data_valid, config['var_equal'], config['is_vs_control'], config['is_non_param'], config['is_paired'])
            if t_name: mtt_test_name = t_name
            min_p = min([p_val for _, _, p_val in pairs]) if pairs else np.nan

        if not np.isnan(min_p) and min_p < 0.05:
            stars = "***" if min_p < 0.001 else "**" if min_p < 0.01 else "*"
            plotted_stars.add(stars)
            max_mean_err_c = max([np.nanmean(d) + calc_error(d, config['error_bar_type']) for d in col_data_valid if not np.isnan(np.nanmean(d)) and not np.isnan(calc_error(d, config['error_bar_type']))] + [0])
            text_y = max_mean_err_c + (mtt_max_y_comb * 0.05)
            mtt_max_y_comb = max(mtt_max_y_comb, text_y * 1.15)
            ax.text(conc_vals_plot[idx_c], text_y, stars, ha='center', va='bottom', fontsize=14, fontweight='bold', color='black')

    if dropped_warnings: st.warning(f"⚠️ データ不足により除外: {', '.join(dropped_warnings)}")
    if non_param_warnings: st.info("💡 n≤3の場合、ノンパラメトリック検定で有意差が出ない可能性があります。")

    ax.set_xscale('log')
    ax.set_ylim(bottom=0, top=mtt_max_y_comb)
    ax.yaxis.set_major_locator(ticker.MultipleLocator(20))
    for spine in ax.spines.values(): spine.set_color('black'); spine.set_linewidth(1.2)
    ax.minorticks_off()
    
    if config['mtt_custom_xticks'].strip() and len(conc_vals_plot) > 0:
        try:
            c_ticks = [float(x.strip()) for x in config['mtt_custom_xticks'].split(',') if x.strip()]
            all_x = conc_vals_plot + c_ticks
            min_x, max_x = min(all_x), max(all_x)
            low_exp, high_exp = int(np.floor(np.log10(min_x))), int(np.ceil(np.log10(max_x)))
            combined_ticks = sorted(list(set([10**e for e in range(low_exp, high_exp + 1)] + c_ticks)))
            ax.set_xticks(combined_ticks)
            ax.get_xaxis().set_major_formatter(ticker.ScalarFormatter())
            ax.set_xlim(min_x * 0.8, max_x * 1.2)
        except Exception: pass
    else: ax.xaxis.set_major_formatter(ticker.FuncFormatter(lambda y, _: '{:g}'.format(y)))
        
    ax.tick_params(direction='in', length=5, width=1.2, labelsize=12, colors='black', which='major')
    ax.set_ylabel(config['ylabel_input'], fontsize=14, fontweight='bold', labelpad=8)
    ax.set_xlabel(f"{config['l_name']} [{config['mtt_unit']}]", fontsize=14, fontweight='bold', labelpad=8)
    if num_p > 1: ax.legend(loc='lower left', frameon=False, prop={'size': 13})
    
    max_n = max([np.count_nonzero(~np.isnan(plates_data[i][valid_rows, c])) for i in range(num_p) for c in s_cols_plot]) if num_p > 0 else 0
    star_str = ", " + ", ".join([f"{s} p < {0.05 if s=='*' else 0.01 if s=='**' else 0.001}" for s in ["*", "**", "***"] if s in plotted_stars]) if plotted_stars else ""
    
    # ★トグルの状態によるタイトル変更
    if config.get('show_stats', True):
        ax.set_title(f"{mtt_test_name}{star_str}, n={max_n}" if mtt_test_name and num_p > 1 else f"n={max_n}", fontsize=14, pad=15, loc='right')
    else:
        ax.set_title(f"n={max_n}", fontsize=14, pad=15, loc='right')

    st.pyplot(fig_comb)
    
    # --- Excel 書き出し ---
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        mtt_summary_dict = {"濃度 (Concentration)": [0.0] + [float(x) for x in conc_vals_plot]}
        err_label = "SEM(%)" if "SEM" in config['error_bar_type'] else "SD(%)"
        for i, p_name in enumerate(plate_names):
            mtt_summary_dict[f"{p_name}_Mean(%)"] = [100.0] + [float(np.nanmean(plates_data[i][valid_rows, c])) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
            mtt_summary_dict[f"{p_name}_{err_label}"] = [float(ctrl_err_pct_list[i])] + [float(calc_error(plates_data[i][valid_rows, c], config['error_bar_type'])) if not np.isnan(plates_data[i][valid_rows, c]).all() else np.nan for c in s_cols_plot]
        pd.DataFrame(mtt_summary_dict).to_excel(writer, sheet_name='Summary', index=False)
        
        ws = writer.book['Summary']
        ws.cell(row=2, column=len(mtt_summary_dict) + 2, value="💡 【エラーバー付き折れ線グラフの最短作成手順】")
        ws.cell(row=3, column=len(mtt_summary_dict) + 2, value="1. 『濃度』の列と、グラフにしたい『〇〇_Mean(%)』の列を同時選択し、[挿入] ＞ [散布図 (直線とマーカー)] を作成。")
        ws.cell(row=4, column=len(mtt_summary_dict) + 2, value="2. 作成されたグラフの横軸をクリックして[軸の書式設定]を開き、『対数目盛を表示する』にチェック。")
        ws.cell(row=5, column=len(mtt_summary_dict) + 2, value="3. グラフのプロット線をクリックし、[＋] ＞ [誤差範囲] ＞ [その他の誤差範囲オプション]。")
        ws.cell(row=6, column=len(mtt_summary_dict) + 2, value="4. 『カスタム』にチェックを入れ、『値の指定』をクリック。")
        ws.cell(row=7, column=len(mtt_summary_dict) + 2, value=f"5. 正負両方の選択ボックスに、対応する『〇〇_{err_label}』の数値をドラッグして指定すれば完成！")

        long_mtt_list = []
        for i, p_name in enumerate(plate_names):
            ctrl_vals = [plates_data[i][r, c] for r in valid_rows for c in c_cols if c not in i_cols]
            for val in ctrl_vals:
                if not np.isnan(val): long_mtt_list.append({"条件名": f"{p_name}_0_{config['mtt_unit']}", "正規化生存率 (%)": float(val)})
            for idx_c, c in enumerate(s_cols_plot):
                for val in plates_data[i][valid_rows, c]:
                    if not np.isnan(val): long_mtt_list.append({"条件名": f"{p_name}_{conc_vals_plot[idx_c]}_{config['mtt_unit']}", "正規化生存率 (%)": float(val)})
        pd.DataFrame(long_mtt_list).to_excel(writer, sheet_name='Normalized_Data', index=False)
        
        for i in range(num_p):
            df_norm = pd.DataFrame(plates_data[i])
            df_norm.index = ['A','B','C','D','E','F','G','H']
            df_norm.columns = [str(x+1) for x in range(12)]
            df_norm.to_excel(writer, sheet_name=re.sub(r'[\\/*?:\[\]]', '', f"Plate_{i+1}_{plate_names[i]}")[:31])

    st.download_button("📥 Excelデータをダウンロード (全データ・統計詳細シート同梱)", excel_buffer.getvalue(), "Analysis_Data.xlsx", type="primary", use_container_width=True)
    dl_col1, dl_col2 = st.columns(2)
    buf_c = io.BytesIO()
    fig_comb.savefig(buf_c, format='svg', bbox_inches='tight')
    with dl_col1: st.download_button("📥 統合グラフ(SVG)を保存", buf_c.getvalue(), "Combined_Graph.svg", "image/svg+xml", use_container_width=True)
    with st.expander("個別プレートのグラフ(SVG)をダウンロード"):
        for p_name, f in indiv_figs:
            buf_i = io.BytesIO()
            f.savefig(buf_i, format='svg', bbox_inches='tight')
            st.download_button(f"📥 {p_name} のグラフ", buf_i.getvalue(), f"{p_name}_Graph.svg", "image/svg+xml")

def render_single_target(input_data, config):
    upper_labels, lower_labels, internal_ids, raw_processed = [], [], [], {}
    dropped_warnings, non_param_warnings = set(), set()
    exclude_flags = [] # ★追加
    
    for idx, item in enumerate(input_data):
        u, d, val_t_list, val_l_list = item[0], item[1], item[2], item[3] if len(item) > 3 else []
        exclude_flag = item[4] if len(item) > 4 else False # ★追加
        exclude_flags.append(exclude_flag)
        
        if config['is_microscope']: raw_processed[f"C_{idx}"] = parse_text(val_t_list[0])
        else:
            t_nums, l_nums = parse_text(val_t_list[0]), parse_text(val_l_list[0])
            length = max(len(t_nums), len(l_nums))
            t_nums.extend([np.nan] * (length - len(t_nums)))
            l_nums.extend([np.nan] * (length - len(l_nums)))
            
            processed = []
            for t, l in zip(t_nums, l_nums):
                if np.isnan(t) or np.isnan(l): processed.append(np.nan)
                elif config['is_qpcr']: processed.append(t - l)
                else: processed.append(np.nan if l == 0 else t / l)
            raw_processed[f"C_{idx}"] = processed
        
        upper_labels.append(u or f"U_{idx+1}")
        lower_labels.append(d or "")
        internal_ids.append(f"C_{idx}")
    
    if not any(len([v for v in raw_processed[uid] if not np.isnan(v)]) > 0 for uid in internal_ids): st.stop()
        
    final_norm = {}
    ctrl_id = internal_ids[0]
    for i, uid in enumerate(internal_ids):
        c_id = internal_ids[lower_labels.index(lower_labels[i])] if 'グループ' in config['norm_mode'] else ctrl_id
        c_mean = np.nanmean(raw_processed[c_id])
        if np.isnan(c_mean) or c_mean == 0: c_mean = 1.0 
        
        if config['is_qpcr']: final_norm[uid] = [2 ** -(v - c_mean) for v in raw_processed[uid]]
        else: final_norm[uid] = [v / c_mean for v in raw_processed[uid]]
    
    p_pairs, test_desc_flat = [], ""
    groupings = [[u for u in internal_ids if lower_labels[internal_ids.index(u)] == low] for low in sorted(list(set(lower_labels)), key=lambda x: lower_labels.index(x))] if config['is_grouped_test'] else [internal_ids]

    for grp in groupings:
        valid_uids, valid_data = [], []
        for u in grp:
            if exclude_flags[internal_ids.index(u)]: continue # ★除外処理
            non_nan_data = [v for v in raw_processed[u] if not np.isnan(v)]
            if len(non_nan_data) >= 2:
                valid_uids.append(u); valid_data.append(non_nan_data)
            else:
                uid_idx = internal_ids.index(u)
                dropped_warnings.add(f"{upper_labels[uid_idx]} ({lower_labels[uid_idx]})" if lower_labels[uid_idx] else upper_labels[uid_idx])
        
        if len(valid_data) < 2: continue
        if config['is_non_param'] and any(len(d) <= 3 for d in valid_data): non_param_warnings.add("解析データ")
        
        # ★トグルONの時だけ検定
        if config.get('show_stats', True):
            _, pairs, t_name = run_statistical_test(valid_data, config['var_equal'], config['is_vs_control'], config['is_non_param'], config['is_paired'])
            if t_name: test_desc_flat = t_name
            for i_idx, j_idx, p_val in pairs: p_pairs.append((valid_uids[i_idx], valid_uids[j_idx], p_val))

    if dropped_warnings: st.warning(f"⚠️ データ不足により除外: {', '.join(dropped_warnings)}")
    if non_param_warnings: st.info("💡 n≤3の場合、ノンパラメトリック検定で有意差が出ない可能性があります。")

    unique_low = sorted(list(set(lower_labels)), key=lambda x: lower_labels.index(x))
    unique_up = sorted(list(set(upper_labels)), key=lambda x: upper_labels.index(x))
    gray_palette = ['black', 'darkgray', 'lightgray', 'dimgray', 'whitesmoke', '#E0E0E0']
    palette = {u: gray_palette[i % len(gray_palette)] for i, u in enumerate(unique_up)} if "色分け" in config['color_mode'] else {u: "black" for u in unique_up}
    
    fig, ax = plt.subplots(figsize=(max(4.0, len(internal_ids)*1.5+1.5), 5.5))
    fig.patch.set_facecolor('white'); ax.set_facecolor('white')
    
    x_coords, bar_width = {}, config['bar_width']
    if config['layout_mode'] == "条件ごとにグループ化":
        current_x = 0
        for low in unique_low:
            members = [i for i, l in enumerate(lower_labels) if l == low]
            for i in members:
                x_coords[internal_ids[i]] = current_x
                current_x += bar_width + 0.02
            current_x += 0.5
    else:
        for i, uid in enumerate(internal_ids): x_coords[uid] = float(i)

    if config['is_microscope']:
        box_data_safe = [[v for v in final_norm[uid] if not np.isnan(v)] or [np.nan] for uid in internal_ids]
        ax.boxplot(box_data_safe, positions=[x_coords[uid] for uid in internal_ids], widths=bar_width*1.5, patch_artist=True, 
                   boxprops=dict(facecolor='white', color='black', linewidth=1.2), 
                   capprops=dict(color='black', linewidth=1.2), whiskerprops=dict(color='black', linewidth=1.2),
                   medianprops=dict(color='black', linewidth=1.5), flierprops=dict(marker='o', markerfacecolor='black', markeredgecolor='black', alpha=0.8, markersize=4))
    else:
        for i, uid in enumerate(internal_ids):
            mean_val, err_val = np.nanmean(final_norm[uid]), calc_error(final_norm[uid], config['error_bar_type'])
            ax.bar(x_coords[uid], mean_val if not np.isnan(mean_val) else 0, yerr=err_val if not np.isnan(err_val) else 0, 
                   width=bar_width, color=palette[upper_labels[i]], edgecolor='black', capsize=3, error_kw=dict(ecolor='black', lw=1.2), 
                   label=upper_labels[i] if i == upper_labels.index(upper_labels[i]) else "")

    for spine in ax.spines.values(): spine.set_visible(True); spine.set_color('black'); spine.set_linewidth(1.5)
    ax.tick_params(axis='y', colors='black', direction='in', left=True, right=False, length=5, width=1.5, labelsize=14)
    ax.tick_params(axis='x', bottom=False, top=False); ax.set_xticklabels([]) 
    trans = transforms.blended_transform_factory(ax.transData, ax.transAxes)
    
    if config['layout_mode'] == "条件ごとにグループ化" and "色分け" in config['color_mode']:
        for low in unique_low:
            xs = [x_coords[internal_ids[i]] for i, l in enumerate(lower_labels) if l == low]
            if xs: ax.text(sum(xs) / len(xs), -0.05, low, ha='center', va='top', transform=trans, fontsize=16, fontweight='bold', color='black')
    else:
        for i, uid in enumerate(internal_ids):
            ax.text(x_coords[uid], -0.05, upper_labels[i], ha='center', va='top', transform=trans, fontsize=16, color='black', fontweight='bold')
        for label, elements in [(k, list(g)) for k, g in itertools.groupby(enumerate(lower_labels), key=lambda x: x[1])]:
            if not label: continue
            xs = [x_coords[internal_ids[x[0]]] for x in elements]
            x_start, x_end = min(xs), max(xs)
            if x_start != x_end: ax.plot([x_start - bar_width/2, x_end + bar_width/2], [-0.16, -0.16], color='black', lw=1.5, transform=trans, clip_on=False)
            ax.text((x_start + x_end) / 2, -0.21, label, ha='center', va='top', transform=trans, fontsize=16, fontweight='bold', color='black')

    all_vals = [v for vals in final_norm.values() for v in vals if not np.isnan(v)]
    current_max_y = max(all_vals + [0]) if all_vals else 1.0
    for uid in internal_ids:
        m, e = np.nanmean(final_norm[uid]), calc_error(final_norm[uid], config['error_bar_type'])
        if not np.isnan(m) and not np.isnan(e): current_max_y = max(current_max_y, m + e)
    if current_max_y == 0: current_max_y = 1.0
    
    y_shift, h, base_bracket_y, max_element_y = current_max_y * 0.15, current_max_y * 0.025, current_max_y * 1.10, current_max_y
    levels, max_level, sig_pairs, plotted_stars = [], 0, [], set()
    
    for u1, u2, p in p_pairs:
        if p >= 0.05 or np.isnan(p) or u1 not in x_coords or u2 not in x_coords: continue
        sig_pairs.append((min(x_coords[u1], x_coords[u2]), max(x_coords[u1], x_coords[u2]), "***" if p < 0.001 else "**" if p < 0.01 else "*"))
    
    for x_start, x_end, stars in sorted(sig_pairs, key=lambda x: x[1] - x[0]):
        plotted_stars.add(stars)
        placed_level = next((l_idx for l_idx, intervals in enumerate(levels) if not any(not (x_end < s or x_start > e) for s, e in intervals)), -1)
        if placed_level == -1: placed_level = len(levels); levels.append([])
        levels[placed_level].append((x_start, x_end)); max_level = max(max_level, placed_level)
        by = base_bracket_y + placed_level * y_shift; max_element_y = max(max_element_y, by + h)
        ax.plot([x_start, x_start, x_end, x_end], [by - h, by, by, by - h], color='black', lw=1.2)
        ax.text((x_start + x_end) / 2, by + h*0.2, stars, ha='center', va='bottom', color='black', fontsize=14, fontweight='bold')

    ax.set_ylim(0, max(current_max_y * 1.2, max_element_y * 1.15))
    x_vals = list(x_coords.values())
    if x_vals: ax.set_xlim(min(x_vals) - 0.6, max(x_vals) + 0.6)
    ax.set_ylabel(config['ylabel_input'], fontsize=16, fontweight="bold", color='black', labelpad=10)
    
    if "色分け" in config['color_mode']:
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        if by_label: ax.legend(by_label.values(), by_label.keys(), loc='upper right', frameon=False, prop={'size': 12, 'weight': 'bold'})

    n_list = [len([v for v in raw_processed[u] if not np.isnan(v)]) for u in internal_ids]
    expected_n = n_list[0] if n_list and len(set(n_list)) == 1 else "varies"
    star_str = ", " + ", ".join([f"{s} p < {0.05 if s=='*' else 0.01 if s=='**' else 0.001}" for s in ["*", "**", "***"] if s in plotted_stars]) if plotted_stars else ""
    title_str = f"{test_desc_flat}{star_str}" if config['is_microscope'] else f"{test_desc_flat}{star_str}, n={expected_n}" if test_desc_flat else f"n={expected_n}"
    
    # ★トグルの状態によるタイトル変更
    if not config.get('show_stats', True): title_str = f"n={expected_n}"
    if title_str: ax.set_title(title_str, fontsize=14, pad=15, loc='right')

    st.pyplot(fig)
    
    # --- Excel 書き出し ---
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        err_label = "SEM" if "SEM" in config['error_bar_type'] else "SD"
        
        summary_df = pd.DataFrame({'上段ラベル': upper_labels, '下段ラベル': lower_labels, '平均': [np.nanmean(final_norm[u]) for u in internal_ids], err_label: [calc_error(final_norm[u], config['error_bar_type']) for u in internal_ids]})
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        pd.DataFrame([{"条件名": f"{upper_labels[i]} ({lower_labels[i]})" if lower_labels[i] else upper_labels[i], "正規化データ": float(val)} for i, u in enumerate(internal_ids) for val in final_norm[u] if not np.isnan(val)]).to_excel(writer, sheet_name='Normalized_Data', index=False)
        
        stat_data = []
        for u1, u2, p in p_pairs:
            idx1, idx2 = internal_ids.index(u1), internal_ids.index(u2)
            c1 = f"{upper_labels[idx1]} ({lower_labels[idx1]})" if lower_labels[idx1] else upper_labels[idx1]
            c2 = f"{upper_labels[idx2]} ({lower_labels[idx2]})" if lower_labels[idx2] else upper_labels[idx2]
            stat_data.append({"比較": f"{c1} vs {c2}", "p値": p if not np.isnan(p) else "N/A", "判定": "***" if p<0.001 else "**" if p<0.01 else "*" if p<0.05 else "ns" if not np.isnan(p) else "N/A"})
        if stat_data: pd.DataFrame(stat_data).to_excel(writer, sheet_name='Statistical_Details', index=False)

        try:
            if config['is_microscope']:
                ws = writer.book['Normalized_Data']
                ws.cell(row=2, column=4, value="💡 【箱ひげ図の最短作成手順】")
                ws.cell(row=3, column=4, value="1. 左のA列とB列をすべて全選択します。")
                ws.cell(row=4, column=4, value="2. [挿入]タブ ＞ [統計グラフ] ＞ [箱ひげ図] をクリックします。")
            elif config['layout_mode'] == "条件ごとにグループ化":
                matrix_mean = pd.DataFrame(index=unique_up, columns=unique_low)
                matrix_sd = pd.DataFrame(index=unique_up, columns=unique_low)
                for i, uid in enumerate(internal_ids):
                    matrix_mean.at[upper_labels[i], lower_labels[i]] = np.nanmean(final_norm[uid])
                    matrix_sd.at[upper_labels[i], lower_labels[i]] = calc_error(final_norm[uid], config['error_bar_type'])
                
                matrix_mean.to_excel(writer, sheet_name='Summary_Matrix', startrow=1, startcol=0)
                matrix_sd.to_excel(writer, sheet_name='Summary_Matrix', startrow=len(unique_up)+4, startcol=0)
                
                ws = writer.book['Summary_Matrix']
                ws.cell(row=1, column=1, value="【平均値 (Mean)】")
                ws.cell(row=len(unique_up)+4, column=1, value=f"【{err_label}】")
                
                sc = len(unique_low) + 3
                ws.cell(row=2, column=sc, value="💡 【グループ化棒グラフの最短作成手順】")
                ws.cell(row=3, column=sc, value="1. 左上の【平均値】の表(A2から)を丸ごと選択し、[挿入] ＞ [2D 縦棒 (集合縦棒)] をクリック。")
                ws.cell(row=4, column=sc, value="2. 追加された棒をクリックし、[誤差範囲] ＞ [その他の誤差範囲オプション] ＞ [カスタム]")
                ws.cell(row=5, column=sc, value=f"3. 値の指定で、下の【{err_label}】の表の該当する行をドラッグして指定すれば完成です！")
            else:
                ws = writer.book['Summary']
                sc = len(summary_df.columns) + 2
                ws.cell(row=2, column=sc, value="💡 【エラーバー付き棒グラフの最短作成手順】")
                ws.cell(row=3, column=sc, value="1. 左の『上段ラベル』と『平均』の列を選択し、[挿入] ＞ [縦棒グラフ] を作成。")
                ws.cell(row=4, column=sc, value="2. グラフの棒をクリックし、[＋] ＞ [誤差範囲] ＞ [その他の誤差範囲オプション]。")
                ws.cell(row=5, column=sc, value="3. 『カスタム』にチェックを入れ、『値の指定』。")
                ws.cell(row=6, column=sc, value=f"4. 正負両方に、左の『{err_label}』の数値をドラッグして指定すれば完成！")
        except Exception:
            pass

    col_dl1, col_dl2 = st.columns(2)
    buf_svg = io.BytesIO()
    fig.savefig(buf_svg, format='svg', bbox_inches='tight')
    with col_dl1: st.download_button("📥 Excelデータをダウンロード", excel_buffer.getvalue(), "Analysis_Data.xlsx", type="primary", use_container_width=True)
    with col_dl2: st.download_button("📥 完成グラフ(SVG)を保存", buf_svg.getvalue(), "Graph.svg", "image/svg+xml", use_container_width=True)


def render_multi_target(input_data, config):
    upper_labels, lower_labels, internal_ids = [], [], []
    raw_processed_multi = {j: {} for j in range(config['num_targets'])}
    dropped_warnings, non_param_warnings = set(), set()
    exclude_flags = [] # ★追加
    
    for idx, item in enumerate(input_data):
        u, d, val_t_list, val_l_list = item[0], item[1], item[2], item[3] if len(item) > 3 else []
        exclude_flag = item[4] if len(item) > 4 else False # ★追加
        exclude_flags.append(exclude_flag)
        
        for j in range(config['num_targets']):
            if config['is_microscope']: raw_processed_multi[j][f"C_{idx}"] = parse_text(val_t_list[j])
            else:
                t_nums, l_nums = parse_text(val_t_list[j]), parse_text(val_l_list[j])
                length = max(len(t_nums), len(l_nums))
                t_nums_ext = t_nums + [np.nan] * (length - len(t_nums))
                l_nums_ext = l_nums + [np.nan] * (length - len(l_nums))
                
                processed = []
                for t, l in zip(t_nums_ext, l_nums_ext):
                    if np.isnan(t) or np.isnan(l): processed.append(np.nan)
                    elif config['is_qpcr']: processed.append(t - l)
                    else: processed.append(np.nan if l == 0 else t / l)
                raw_processed_multi[j][f"C_{idx}"] = processed
                
        upper_labels.append(u or f"U_{idx+1}")
        lower_labels.append(d or "")
        internal_ids.append(f"C_{idx}")
        
    if not any(len([v for v in raw_processed_multi[0][uid] if not np.isnan(v)]) > 0 for uid in internal_ids): st.stop()
    
    final_norm_multi = {j: {} for j in range(config['num_targets'])}
    ctrl_id = internal_ids[0]
    
    for j in range(config['num_targets']):
        for i, uid in enumerate(internal_ids):
            c_id = internal_ids[lower_labels.index(lower_labels[i])] if 'グループ' in config['norm_mode'] else ctrl_id
            c_mean = np.nanmean(raw_processed_multi[j][c_id])
            if np.isnan(c_mean) or c_mean == 0: c_mean = 1.0 
            if config['is_qpcr']: final_norm_multi[j][uid] = [2 ** -(v - c_mean) for v in raw_processed_multi[j][uid]]
            else: final_norm_multi[j][uid] = [v / c_mean for v in raw_processed_multi[j][uid]]
    
    p_pairs_multi = {j: [] for j in range(config['num_targets'])}
    test_desc_flat = ""
    groupings = [[u for u in internal_ids if lower_labels[internal_ids.index(u)] == low] for low in sorted(list(set(lower_labels)), key=lambda x: lower_labels.index(x))] if config['is_grouped_test'] else [internal_ids]

    for j in range(config['num_targets']):
        for grp in groupings:
            valid_uids, valid_data = [], []
            for u in grp:
                if exclude_flags[internal_ids.index(u)]: continue # ★除外処理
                non_nan_data = [v for v in raw_processed_multi[j][u] if not np.isnan(v)]
                if len(non_nan_data) >= 2:
                    valid_uids.append(u); valid_data.append(non_nan_data)
                else:
                    uid_idx = internal_ids.index(u)
                    dropped_warnings.add(f"{config['target_names'][j]}の{upper_labels[uid_idx]}")
            
            if len(valid_data) < 2: continue
            if config['is_non_param'] and any(len(d) <= 3 for d in valid_data): non_param_warnings.add("解析データ")
                
            # ★トグルONの時だけ検定
            if config.get('show_stats', True):
                _, pairs, t_name = run_statistical_test(valid_data, config['var_equal'], config['is_vs_control'], config['is_non_param'], config['is_paired'])
                if t_name: test_desc_flat = t_name
                for i_idx, j_idx, p_val in pairs: p_pairs_multi[j].append((valid_uids[i_idx], valid_uids[j_idx], p_val))

    if dropped_warnings: st.warning(f"⚠️ データ不足により除外: {', '.join(dropped_warnings)}")
    if non_param_warnings: st.info("💡 n≤3の場合、ノンパラメトリック検定で有意差が出ない可能性があります。")

    unique_up = sorted(list(set(upper_labels)), key=lambda x: upper_labels.index(x))
    gray_palette = ['black', 'darkgray', 'lightgray', 'dimgray', 'whitesmoke', '#E0E0E0']
    if "色分け" in config['color_mode']:
        colors = sns.color_palette("Set1", max(len(unique_up), 2))
        palette = {u: colors[i % len(colors)] for i, u in enumerate(unique_up)}
    else:
        palette = {u: gray_palette[i % len(gray_palette)] for i, u in enumerate(unique_up)}
        
    fig, ax = plt.subplots(figsize=(max(6.0, config['num_targets'] * len(unique_up) * 1.0), 5.5))
    fig.patch.set_facecolor('white'); ax.set_facecolor('white')
    
    bar_width, x_coords_multi, target_centers, current_x = config['bar_width'], {j: {} for j in range(config['num_targets'])}, [], 0
    
    for j in range(config['num_targets']):
        g_start = current_x
        for i, uid in enumerate(internal_ids):
            x_coords_multi[j][uid] = current_x
            if config['is_microscope']:
                d_list_safe = [v for v in final_norm_multi[j][uid] if not np.isnan(v)] or [np.nan]
                ax.boxplot([d_list_safe], positions=[current_x], widths=bar_width*1.5, patch_artist=True, 
                           boxprops=dict(facecolor='white', color='black', linewidth=1.2), capprops=dict(color='black', linewidth=1.2),
                           whiskerprops=dict(color='black', linewidth=1.2), medianprops=dict(color='black', linewidth=1.5), 
                           flierprops=dict(marker='o', markerfacecolor='black', markeredgecolor='black', alpha=0.8, markersize=4))
            else:
                mean_val, err_val = np.nanmean(final_norm_multi[j][uid]), calc_error(final_norm_multi[j][uid], config['error_bar_type'])
                ax.bar(current_x, mean_val if not np.isnan(mean_val) else 0, yerr=err_val if not np.isnan(err_val) else 0, 
                       width=bar_width, color=palette[upper_labels[i]], edgecolor='black', capsize=3, error_kw=dict(ecolor='black', lw=1.2), 
                       label=upper_labels[i] if (j == 0 and i == upper_labels.index(upper_labels[i])) else "")
            current_x += bar_width + 0.02
        target_centers.append((g_start + current_x - bar_width - 0.02) / 2)
        current_x += 0.8

    for spine in ax.spines.values(): spine.set_visible(True); spine.set_color('black'); spine.set_linewidth(1.5)
    ax.tick_params(axis='y', colors='black', direction='in', left=True, right=False, length=5, width=1.5, labelsize=14)
    ax.tick_params(axis='x', bottom=False, top=False)
    ax.set_xticks(target_centers); ax.set_xticklabels(config['target_names'], fontsize=16, fontweight='bold', color='black')
    
    all_vals = [v for j in range(config['num_targets']) for vals in final_norm_multi[j].values() for v in vals if not np.isnan(v)]
    current_max_y = max(all_vals + [0]) if all_vals else 1.0
    for j in range(config['num_targets']):
        for uid in internal_ids:
            m, e = np.nanmean(final_norm_multi[j][uid]), calc_error(final_norm_multi[j][uid], config['error_bar_type'])
            if not np.isnan(m) and not np.isnan(e): current_max_y = max(current_max_y, m + e)
    if current_max_y == 0: current_max_y = 1.0
    
    y_shift, h, base_bracket_y, max_element_y = current_max_y * 0.15, current_max_y * 0.025, current_max_y * 1.10, current_max_y
    plotted_stars = set()
    
    for j in range(config['num_targets']):
        levels, max_level, sig_pairs = [], 0, []
        for u1, u2, p in p_pairs_multi[j]:
            if p >= 0.05 or np.isnan(p) or u1 not in x_coords_multi[j] or u2 not in x_coords_multi[j]: continue
            sig_pairs.append((min(x_coords_multi[j][u1], x_coords_multi[j][u2]), max(x_coords_multi[j][u1], x_coords_multi[j][u2]), "***" if p < 0.001 else "**" if p < 0.01 else "*"))
        
        for x_start, x_end, stars in sorted(sig_pairs, key=lambda x: x[1] - x[0]):
            plotted_stars.add(stars)
            placed_level = next((l_idx for l_idx, intervals in enumerate(levels) if not any(not (x_end < s or x_start > e) for s, e in intervals)), -1)
            if placed_level == -1: placed_level = len(levels); levels.append([])
            levels[placed_level].append((x_start, x_end)); max_level = max(max_level, placed_level)
            by = base_bracket_y + placed_level * y_shift; max_element_y = max(max_element_y, by + h)
            ax.plot([x_start, x_start, x_end, x_end], [by - h, by, by, by - h], color='black', lw=1.2)
            ax.text((x_start + x_end) / 2, by + h*0.2, stars, ha='center', va='bottom', color='black', fontsize=14, fontweight='bold')
        base_bracket_y += (max_level + 1) * y_shift if sig_pairs else 0
        max_element_y = max(max_element_y, base_bracket_y)

    ax.set_ylim(0, max(current_max_y * 1.2, max_element_y * 1.15))
    ax.set_xlim(-0.6, current_x - 0.8 + 0.6)
    ax.set_ylabel(config['ylabel_input'], fontsize=16, fontweight="bold", color='black', labelpad=10)
    
    if "色分け" in config['color_mode']:
        handles, labels = ax.get_legend_handles_labels()
        by_label = dict(zip(labels, handles))
        if by_label: ax.legend(by_label.values(), by_label.keys(), loc='upper right', frameon=False, prop={'size': 12, 'weight': 'bold'})

    expected_n = n_list[0] if (n_list := [len([v for v in raw_processed_multi[0][u] if not np.isnan(v)]) for u in internal_ids]) and len(set(n_list)) == 1 else "varies"
    star_str = ", " + ", ".join([f"{s} p < {0.05 if s=='*' else 0.01 if s=='**' else 0.001}" for s in ["*", "**", "***"] if s in plotted_stars]) if plotted_stars else ""
    title_str = f"{test_desc_flat}{star_str}, n={expected_n}" if test_desc_flat else f"n={expected_n}"
    
    # ★トグルの状態によるタイトル変更
    if not config.get('show_stats', True): title_str = f"n={expected_n}"
    ax.set_title(title_str, fontsize=14, pad=15, loc='right')

    st.pyplot(fig)
    
    # --- Excel 書き出し ---
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        err_label = "SEM" if "SEM" in config['error_bar_type'] else "SD"
        pd.DataFrame([{'ターゲット名': config['target_names'][j], '上段ラベル': upper_labels[i], '下段ラベル': lower_labels[i], '平均': np.nanmean(final_norm_multi[j][u]), err_label: calc_error(final_norm_multi[j][u], config['error_bar_type'])} for j in range(config['num_targets']) for i, u in enumerate(internal_ids)]).to_excel(writer, sheet_name='Summary', index=False)
        
        ws = writer.book['Summary']
        ws.cell(row=2, column=7, value="💡 【複数ターゲットの棒グラフ最短作成手順】")
        ws.cell(row=3, column=7, value="1. 1行目（見出し）を選択し、[データ]タブ ＞ [フィルター] をクリック。")
        ws.cell(row=4, column=7, value="2. 『ターゲット名』の▼をクリックし、グラフにしたいターゲットを1つだけ選んで絞り込む。")
        ws.cell(row=5, column=7, value="3. 表示された『上段ラベル』と『平均』の列を同時選択し、[挿入] ＞ [縦棒グラフ] を作成。")
        ws.cell(row=6, column=7, value="4. グラフの棒をクリックし、[＋] ＞ [誤差範囲] ＞ [その他の誤差範囲オプション]。")
        ws.cell(row=7, column=7, value=f"5. 『カスタム』を選び、『値の指定』で正負両方に、フィルター後の表示されている『{err_label}』の列をドラッグして指定すれば完成！")

        pd.DataFrame([{"ターゲット名": config['target_names'][j], "条件名": f"{upper_labels[i]} ({lower_labels[i]})" if lower_labels[i] else upper_labels[i], "正規化データ": float(val)} for j in range(config['num_targets']) for i, u in enumerate(internal_ids) for val in final_norm_multi[j][u] if not np.isnan(val)]).to_excel(writer, sheet_name='Normalized_Data', index=False)
        
        stat_data = []
        for j in range(config['num_targets']):
            for u1, u2, p in p_pairs_multi[j]:
                idx1, idx2 = internal_ids.index(u1), internal_ids.index(u2)
                c1 = f"{upper_labels[idx1]} ({lower_labels[idx1]})" if lower_labels[idx1] else upper_labels[idx1]
                c2 = f"{upper_labels[idx2]} ({lower_labels[idx2]})" if lower_labels[idx2] else upper_labels[idx2]
                stat_data.append({"ターゲット名": config['target_names'][j], "比較": f"{c1} vs {c2}", "p値": p if not np.isnan(p) else "N/A", "判定": "***" if p<0.001 else "**" if p<0.01 else "*" if p<0.05 else "ns" if not np.isnan(p) else "N/A"})
        if stat_data: pd.DataFrame(stat_data).to_excel(writer, sheet_name='Statistical_Details', index=False)

    col_dl1, col_dl2 = st.columns(2)
    buf_svg = io.BytesIO()
    fig.savefig(buf_svg, format='svg', bbox_inches='tight')
    with col_dl1: st.download_button("📥 Excelデータをダウンロード", excel_buffer.getvalue(), "Analysis_Data.xlsx", type="primary", use_container_width=True)
    with col_dl2: st.download_button("📥 完成グラフ(SVG)を保存", buf_svg.getvalue(), "Graph.svg", "image/svg+xml", use_container_width=True)
